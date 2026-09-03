# -*- coding: utf-8 -*-
import os as _os
import sys as _sys
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
"""验证 Excel/CSV/TXT 导入：表头识别、标题行、单位行、编码、裁剪"""
import os
import tempfile
import tkinter as tk
from tkinter import messagebox

import numpy as np

for _fn in ('showinfo', 'showwarning', 'showerror', 'askyesno'):
    setattr(messagebox, _fn, lambda *a, **k: True)

from src.core import ship_app_actions
for _fn in ('showinfo', 'showwarning', 'showerror', 'askyesno'):
    setattr(ship_app_actions.messagebox, _fn, lambda *a, **k: True)

root = tk.Tk()
root.withdraw()
from src.app.ship_app import ShipApp
app = ShipApp(root, icon_dir='icon')
root.update()

tmp = tempfile.mkdtemp()
ok = True


def check(name, cond, detail=''):
    global ok
    print('[%s] %s %s' % ('PASS' if cond else 'FAIL', name, detail))
    if not cond:
        ok = False


def load(path):
    headers, rows = app._read_table_file(path)
    rows = app.fill_merged_cells(rows)
    return app.extract_header_and_data(headers, rows)


# ================= 1. 普通 xlsx（首行即表头） =================
print('=' * 60)
print('1. 普通 xlsx')
print('=' * 60)
from openpyxl import Workbook
p1 = os.path.join(tmp, 'plain.xlsx')
wb = Workbook()
ws = wb.active
ws.append(['站号', '半宽', '高度', '系数'])
ws.append([0, 0, 0, 1])
ws.append([1, 50, 2, 1])
ws.append([2, 80, 4, 1])
wb.save(p1)
h1, d1 = load(p1)
check('表头正确', h1 == ['站号', '半宽', '高度', '系数'], str(h1))
check('数据 3 行', len(d1) == 3, str(len(d1)))
check('数值已转换', d1[1][1] == 50.0, str(d1[1]))

# ================= 2. 标题行（合并单元格）+ 表头 =================
print()
print('=' * 60)
print('2. 标题行 + 表头')
print('=' * 60)
p2 = os.path.join(tmp, 'title.xlsx')
wb = Workbook()
ws = wb.active
ws.merge_cells('A1:D1')
ws['A1'] = '某船型值表'
ws.append(['站号', '半宽', '高度', '系数'])
ws.append([0, 0, 0, 1])
ws.append([1, 50, 2, 1])
wb.save(p2)
h2, d2 = load(p2)
check('跳过标题行，表头正确', h2 == ['站号', '半宽', '高度', '系数'], str(h2))
check('数据 2 行（不含标题行）', len(d2) == 2, str(len(d2)))
check('列数未因标题行被截断（4 列）', len(h2) == 4, str(len(h2)))

# ================= 3. 单位行（表头下方 m m m） =================
print()
print('=' * 60)
print('3. 单位行')
print('=' * 60)
p3 = os.path.join(tmp, 'units.xlsx')
wb = Workbook()
ws = wb.active
ws.append(['站号', '半宽', '高度', '系数'])
ws.append(['m', 'm', 'm', ''])
ws.append([0, 0, 0, 1])
ws.append([1, 50, 2, 1])
ws.append([2, 80, 4, 1])
wb.save(p3)
h3, d3 = load(p3)
check('表头正确', h3 == ['站号', '半宽', '高度', '系数'], str(h3))
check('单位行被丢弃，数据 3 行', len(d3) == 3, str(len(d3)))
check('单位行内容未混入数据', d3[0][0] == 0.0, str(d3[0]))

# ================= 4. GBK 编码 CSV =================
print()
print('=' * 60)
print('4. GBK 编码 CSV')
print('=' * 60)
p4 = os.path.join(tmp, 'gbk.csv')
with open(p4, 'w', encoding='gbk', newline='') as f:
    f.write('站号,半宽,高度\n')
    f.write('0,0,0\n')
    f.write('1,50,2\n')
    f.write('2,80,4\n')
h4, d4 = load(p4)
check('GBK 表头正确', h4 == ['站号', '半宽', '高度'], str(h4))
check('GBK 数据 3 行', len(d4) == 3, str(len(d4)))

# ================= 5. 尾部空行/空列裁剪 =================
print()
print('=' * 60)
print('5. 尾部空行/空列裁剪')
print('=' * 60)
p5 = os.path.join(tmp, 'trail.xlsx')
wb = Workbook()
ws = wb.active
ws.append(['站号', '半宽', '高度'])
ws.append([0, 0, 0])
ws.append([1, 50, 2])
ws.append([None, None, None])   # 空行（需样式单元格才可能出现在 xlsx 中）
ws.append([None, None, None])
wb.save(p5)
h5, d5 = load(p5)
check('尾部空行被裁剪，数据 2 行', len(d5) == 2, str(len(d5)))

# ================= 6. .xls 缺少 xlrd 时给出明确提示 =================
print()
print('=' * 60)
print('6. .xls 依赖检查')
print('=' * 60)
p6 = os.path.join(tmp, 'old.xls')
with open(p6, 'wb') as f:
    f.write(b'D0CF11E0A1B11AE1\x00\x00')
try:
    app._read_table_file(p6)
    check('.xls 缺失 xlrd 时抛异常', False, '未抛出异常')
except RuntimeError as e:
    check('.xls 缺失 xlrd 时抛异常', 'xlrd' in str(e) or '.xls' in str(e), str(e)[:60])

# ================= 7. TXT 制表符分隔 =================
print()
print('=' * 60)
print('7. TXT 制表符分隔')
print('=' * 60)
p7 = os.path.join(tmp, 'tab.txt')
with open(p7, 'w', encoding='utf-8') as f:
    f.write('站号\t半宽\t高度\n')
    f.write('0\t0\t0\n')
    f.write('1\t50\t2\n')
h7, d7 = load(p7)
check('TXT 表头正确', h7 == ['站号', '半宽', '高度'], str(h7))
check('TXT 数据 2 行', len(d7) == 2, str(len(d7)))

print()
print('IMPORT EXCEL TEST %s' % ('PASS' if ok else 'FAIL'))
root.destroy()
