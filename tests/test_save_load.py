# -*- coding: utf-8 -*-
import os as _os
import sys as _sys
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
"""验证项目保存/加载机制：原子写入、回读校验、版本化、树重建分组、脏标记、导出"""
import os
import tempfile
import tkinter as tk
from tkinter import messagebox

import numpy as np

for _fn in ('showinfo', 'showwarning', 'showerror', 'askyesno', 'askyesnocancel'):
    setattr(messagebox, _fn, lambda *a, **k: True)

from src.core import ship_app_actions
for _fn in ('showinfo', 'showwarning', 'showerror', 'askyesno', 'askyesnocancel'):
    setattr(ship_app_actions.messagebox, _fn, lambda *a, **k: True)
ship_app_actions.ask_text_dialog = lambda *a, **k: ''
ship_app_actions.ask_numeric_dialog = lambda *a, **k: None
ship_app_actions.ask_multi_select = lambda *a, **k: [1]
ship_app_actions.ask_multiline_input = lambda *a, **k: ''

root = tk.Tk()
root.withdraw()
from src.app.ship_app import ShipApp
app = ShipApp(root, icon_dir='icon')
root.update()

tmp = tempfile.mkdtemp()
ok = True


def check(name, cond, detail=''):
    global ok
    print('[%s] %s %s' % ('PASS' if cond else 'FAIL', name, detail))
    if not cond:
        ok = False


# ================= 准备一组状态 =================
app.Lpp, app.Breadth, app.Depth = 100.0, 12.0, 8.0
app.LppStartStation, app.LppEndStation = 0.0, 10.0
app.Half_table.set_columns(['列', '站号', '半宽', '系数', '相对矩臂'])
app.Half_table.set_data([[1, 0, 0, 1, 0], [2, 5, 6, 1, 0], [3, 10, 0, 1, 0]])
app.Z_table.set_columns(['列', '高度', '半宽', '系数'])
app.Z_table.set_data([[1, 0, 0, 1], [2, 2, 4, 1], [3, 4, 5, 1]])
app.waterlines = [{'type': 'waterline', 'name': 'WL0', 'height': 0.0,
                   'table': {'columns': ['列', '站号', '半宽'], 'rows': [[1, 0, 0], [2, 5, 6], [3, 10, 0]]}},
                  {'type': 'waterline', 'name': 'WL2', 'height': 2.0,
                   'table': {'columns': ['列', '站号', '半宽'], 'rows': [[1, 0, 0], [2, 5, 5], [3, 10, 0]]}}]
app.bodyplans = [{'type': 'bodyplan', 'name': '站 0', 'station': 0.0,
                  'table': {'columns': ['列', '高度', '半宽'], 'rows': [[1, 0, 0], [2, 2, 1]]}},
                 {'type': 'bodyplan', 'name': '站 5', 'station': 5.0,
                  'table': {'columns': ['列', '高度', '半宽'], 'rows': [[1, 0, 0], [2, 2, 5]]}}]
app.sections = {0.0: {'Y': [0.0, 1.0], 'Z': [0.0, 2.0]},
                5.0: {'Y': [0.0, 5.0], 'Z': [0.0, 2.0]}}
app.IsSymmetricView = True
app.WireframeMode = '高光边缘'
app.original_headers = ['station', 'half_wl']
app.original_data = [[0, 0], [5, 6], [10, 0]]
# 轻量可序列化 ML 模型
from sklearn.tree import DecisionTreeClassifier
app.ML_model = {'model': DecisionTreeClassifier().fit(
    np.array([[0.0], [1.0]]), np.array(['a', 'b'])),
    'required_variables': ['f1'], 'kind': 'Python', 'path': None}

# ================= 1. 保存项目（原子 + 回读校验） =================
print('=' * 60)
print('1. 保存项目')
print('=' * 60)
scs_path = os.path.join(tmp, 'proj.scs')
ship_app_actions.filedialog.asksaveasfilename = lambda **k: scs_path
app._dirty = True
app.menu_save_project()
import pickle
check('文件已生成', os.path.exists(scs_path) and os.path.getsize(scs_path) > 0,
      'size=%d' % os.path.getsize(scs_path) if os.path.exists(scs_path) else '0')
with open(scs_path, 'rb') as f:
    p = pickle.load(f)
check('payload 版本合法 (2/3)', p.get('version') in (2, 3), 'version=%s' % p.get('version'))
check('主尺度已保存', p['principal']['Lpp'] == 100.0)
check('水线面已保存', len(p['waterlines']) == 2)
check('横剖面已保存', len(p['bodyplans']) == 2)
check('sections 已保存', len(p['sections']) == 2)
check('ML 模型已保存', p.get('ml_model') is not None and 'model' in p['ml_model'])
check('UI 状态已保存', p.get('ui_state', {}).get('WireframeMode') == '高光边缘')
check('保存后清除脏标记', app._dirty is False)
check('保存后窗口标题不带 *', not str(root.title()).endswith(' *'))

# 原子写：目录中不应残留 .tmp
leftover = [f for f in os.listdir(tmp) if f.endswith('.tmp')]
check('无临时文件残留', not leftover, str(leftover))

# ================= 2. 脏标记 =================
print()
print('=' * 60)
print('2. 脏标记')
print('=' * 60)
app._clear_dirty()
app.Half_table._push_undo()
app.Half_table.set_cell(0, 1, 999.0)
app._on_half_table_edit(0, 1, 1.0, 999.0)   # 变更前回调 → _save_undo → 标记
check('表格编辑后标记未保存', app._dirty is True)
app._clear_dirty()
app._mark_dirty()
check('任意修改可标记', app._dirty is True)

# ================= 3. 加载项目（还原 + 树重建分组） =================
print()
print('=' * 60)
print('3. 加载项目与树重建')
print('=' * 60)
# 重置状态
app.Lpp = 0.0
app.waterlines = []
app.bodyplans = []
app.sections = {}
app.ML_model = None
app._dirty = True
ship_app_actions.filedialog.askopenfilename = lambda **k: scs_path
app.menu_import_project()
check('加载后主尺度还原', app.Lpp == 100.0)
check('加载后水线面还原', len(app.waterlines) == 2)
check('加载后 ML 模型还原', app.ML_model is not None and app.ML_model.get('kind') == 'Python')
check('加载后 UI 状态还原', app.WireframeMode == '高光边缘' and app.IsSymmetricView is True)
check('加载后清除脏标记', app._dirty is False)

# 树重建分组：所有水线面在同一"船型模型"节点下
ships = [c for c in app.tree.get_children(app.face_root)
         if app.tree.item(c, 'text') == '船型模型']
check('Face 下只有一个"船型模型"节点', len(ships) == 1, '实际 %d' % len(ships))
if ships:
    halves = [c for c in app.tree.get_children(ships[0])
              if app.tree.item(c, 'text') == 'Half']
    check('船型模型下有一个 Half 节点', len(halves) == 1)
    if halves:
        check('Half 下有 2 条水线面',
              len(app.tree.get_children(halves[0])) == 2,
              '实际 %d' % len(app.tree.get_children(halves[0])))
    bodies = [c for c in app.tree.get_children(ships[0])
              if app.tree.item(c, 'text') == 'Body Plan']
    check('船型模型下有一个 Body Plan 节点', len(bodies) == 1)

# 原表格节点已重建
table_children = [app.tree_meta.get(c, {}).get('type', '') for c in app.tree.get_children(app.table_root)]
check('原表格节点已重建', 'table' in table_children, str(table_children))

# ================= 4. 导出表格 =================
print()
print('=' * 60)
print('4. 导出表格')
print('=' * 60)
csv_path = os.path.join(tmp, 'out.csv')
ship_app_actions.filedialog.asksaveasfilename = lambda **k: csv_path
app.isTap = 2  # 半宽
app.menu_export()
with open(csv_path, 'r', encoding='utf-8-sig') as f:
    content = f.read()
check('CSV 导出成功且含表头', '站号' in content and '2,5,6' in content, content[:80])

xlsx_path = os.path.join(tmp, 'out.xlsx')
ship_app_actions.filedialog.asksaveasfilename = lambda **k: xlsx_path
app.menu_export()
from openpyxl import load_workbook
wb2 = load_workbook(xlsx_path, data_only=True)
check('xlsx 导出多工作表', '半宽表' in wb2.sheetnames and '横剖面表' in wb2.sheetnames,
      str(wb2.sheetnames))

print()
print('SAVE/LOAD TEST %s' % ('PASS' if ok else 'FAIL'))
root.destroy()
