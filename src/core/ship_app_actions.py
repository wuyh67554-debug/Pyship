# -*- coding: utf-8 -*-
"""
ship_app_actions.py —— 业务逻辑（导入/导出/机器学习/树/水线面/横剖面操作）
"""

import os
import math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import numpy as np

from src.core import ship_core as core
from src.core import ml_utils
from src.ui.ui_widgets import (EditableTable, ask_choice_dialog, ask_multiline_input,
                        ask_numeric_dialog, ask_multi_select, ask_text_dialog)


class ShipAppActions:
    # =====================================================================
    # 表格编辑回调
    # =====================================================================
    def _on_half_table_edit(self, row, col, old, new):
        """变更前回调：保存撤销快照（此时数据尚未被修改）"""
        self._save_undo('Half_table', '编辑单元格')

    def _on_half_table_plot(self, row, col, old, new):
        """变更后回调：刷新水线面图"""
        self.update_half_width_plot()

    def _on_z_table_edit(self, row, col, old, new):
        """变更前回调：保存撤销快照（此时数据尚未被修改）"""
        self._save_undo('Z_table', '编辑单元格')

    def _on_z_table_plot(self, row, col, old, new):
        """变更后回调：刷新横剖面图"""
        self.update_transverse_section_plot()

    # =====================================================================
    # 数据导入
    # =====================================================================
    def import_table_clicked(self):
        self.clear_logs()
        path = filedialog.askopenfilename(
            title='请选择一个型值表文件',
            filetypes=[('表格文件', '*.xlsx;*.xls;*.csv;*.txt'), ('所有文件', '*.*')],
            parent=self.root)
        if not path:
            self.log('用户取消了文件选择。')
            return
        self.log('开始导入文件: %s' % path)
        try:
            headers, rows = self._read_table_file(path)
        except Exception as e:
            messagebox.showerror('处理失败', '读取文件失败：%s' % e, parent=self.root)
            return
        self.log('文件读取完毕，原始数据为 %d 行 × %d 列。' % (len(rows), len(headers)))

        # 向下填充（合并单元格）
        rows = self.fill_merged_cells(rows)
        # 智能识别表头
        headers, data = self.extract_header_and_data(headers, rows)
        if not data:
            messagebox.showerror(
                '数据为空',
                '识别出的表头下方没有数据行。\n'
                '请检查文件内容（如确认表头与数据在同一工作表中）。',
                parent=self.root)
            return
        self.RawDataStorage = data
        self.log('表头识别完成：%d 列；数据 %d 行。' % (len(headers), len(data)))

        # 清洗单元格
        cleaned = []
        for row in data:
            cleaned.append(['' if v is None or (isinstance(v, float) and math.isnan(v)) else v
                            for v in row])

        # 创建 Table 节点
        table_node = self._tree_add(self.table_root, os.path.basename(path),
                                    {'type': 'table', 'Data': cleaned, 'Headers': headers,
                                     'VariableNames': headers})
        self.original_headers = list(headers)
        self.original_data = cleaned
        self.original_table.set_columns(headers)
        self.original_table.set_data(cleaned)
        self.log('表格 "%s" 的节点已创建。' % os.path.basename(path))
        self.log('文件导入并处理成功！')
        self._mark_dirty()
        messagebox.showinfo('操作成功', '文件导入并处理成功！', parent=self.root)

    @staticmethod
    def _try_convert(value):
        """
        兼容 MATLAB readcell：将数字样式的字符串转为 float，
        无法转换的保持字符串。
        """
        if isinstance(value, str):
            s = value.strip()
            if s != '':
                try:
                    f = float(s)
                    return f
                except ValueError:
                    return value
            return value
        if value is None:
            return ''
        return value

    @staticmethod
    def _read_with_encodings(path, reader):
        """按 utf-8-sig → gbk → gb18030 → latin-1 依次尝试解码文本文件。

        Excel 另存的 CSV 常见为 GBK/ANSI 编码，仅 utf-8 会解不出来。
        """
        last_err = None
        for enc in ('utf-8-sig', 'gbk', 'gb18030', 'latin-1'):
            try:
                with open(path, 'r', encoding=enc, newline='') as f:
                    return reader(f)
            except UnicodeDecodeError as e:
                last_err = e
        if last_err is not None:
            raise last_err
        raise RuntimeError('无法解码文件：%s' % path)

    def _read_excel_rows(self, path):
        """读取 .xlsx / .xls 的全部行（已转值，去掉尾部全空行/全空列）"""
        ext = os.path.splitext(path)[1].lower()
        if ext == '.xls':
            # openpyxl 不支持旧版二进制 .xls，需 xlrd + pandas
            try:
                import xlrd  # noqa: F401
            except ImportError:
                raise RuntimeError(
                    '当前环境未安装 xlrd，无法读取旧版 .xls 二进制格式。\n'
                    '请执行 pip install xlrd\n'
                    '或直接用 Excel/WPS 将文件另存为 .xlsx / .csv 后再导入。')
            try:
                import pandas as pd
            except ImportError:
                raise RuntimeError('读取 .xls 需要 pandas 支持。')
            try:
                df = pd.read_excel(path, dtype=object, header=None, sheet_name=0)
                rows = df.where(pd.notna(df), None).values.tolist()
            except Exception as e:
                raise RuntimeError('读取 .xls 失败：%s\n建议另存为 .xlsx 或 .csv 后导入。' % e)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        rows = [[self._try_convert(v) for v in r] for r in rows]
        return self._trim_empty(rows)

    def _read_csv_rows(self, path):
        import csv

        def _load(f):
            return [r for r in csv.reader(f) if any(str(c).strip() for c in r)]

        rows = self._read_with_encodings(path, _load)
        rows = [[self._try_convert(c) for c in r] for r in rows]
        return self._trim_empty(rows)

    def _read_txt_rows(self, path):
        """txt：按首个非空行判断分隔符（制表符优先，其次逗号），整体用同一分隔符"""
        import csv

        def _load(f):
            lines = [l.rstrip('\r\n') for l in f]
            has_tab = any('\t' in l for l in lines)
            has_comma = any(',' in l for l in lines)
            if not has_tab and not has_comma:
                return [l.split() for l in lines if l.strip()]
            delim = '\t' if has_tab else ','
            return [[c.strip() for c in l.split(delim)] for l in lines
                    if any(ch.strip() for ch in l.split(delim))]

        rows = self._read_with_encodings(path, _load)
        rows = [[self._try_convert(c) for c in r] for r in rows]
        return self._trim_empty(rows)

    @staticmethod
    def _trim_empty(rows):
        """去掉尾部全空行；去掉右侧全空列"""
        while rows and all(v is None or v == '' for v in rows[-1]):
            rows.pop()
        if rows:
            ncol = max(len(r) for r in rows)
            while ncol > 0 and all(len(r) < ncol or r[ncol - 1] in (None, '')
                                   for r in rows):
                ncol -= 1
            rows = [r[:ncol] for r in rows]
        return rows

    def _read_table_file(self, path):
        """读取表格文件，返回 (headers, rows)

        headers 为首行（尽力而为的猜测），rows 为首行之后的数据行；
        更可靠的表头定位由 extract_header_and_data 完成。
        """
        ext = os.path.splitext(path)[1].lower()
        if ext in ('.xlsx', '.xls'):
            rows = self._read_excel_rows(path)
        elif ext == '.csv':
            rows = self._read_csv_rows(path)
        else:
            rows = self._read_txt_rows(path)
        if not rows:
            return ['列'], []
        headers = [self._hdr_str(x) if x is not None and str(x).strip() != ''
                   else 'UnnamedColumn_%d' % (i + 1)
                   for i, x in enumerate(rows[0])]
        return headers, rows[1:]

    @staticmethod
    def _hdr_str(v):
        """表头单元格转字符串：整数型数值不带小数点"""
        if isinstance(v, (float, np.floating)) and math.isfinite(v) and v.is_integer():
            return str(int(v))
        return str(v)

    def fill_merged_cells(self, rows):
        """向下填充（MATLAB fillMergedCells）"""
        self.log('开始执行"向下填充"以处理合并单元格...')
        if not rows:
            return rows
        n_rows = len(rows)
        n_cols = max(len(r) for r in rows)
        filled = [list(r) + [''] * (n_cols - len(r)) for r in rows]
        for c in range(n_cols):
            for r in range(1, n_rows):
                cur = filled[r][c]
                if cur is None or cur == '' or (isinstance(cur, float) and math.isnan(cur)):
                    filled[r][c] = filled[r - 1][c]
        self.log('"向下填充"处理完毕。')
        return filled

    def extract_header_and_data(self, headers, rows):
        """智能识别表头行并切分数据，返回 (headers, data)。

        兼容常见船舶型值表结构：
          · 首行即表头（无标题行）
          · 首行为标题（合并单元格"XX船型值表"），下一行为表头
          · 表头下方有单位行（如 m / m / m）
        n_cols 取所有行的最大宽度，避免标题行只有 1 列导致列数被截断。

        表头候选行评分（取前 min(20, n) 行中得分最高者）：
          · 非空单元格占比（至少 2 个非空）
          · 文本占比（表头几乎全为文本；纯数值行排除）
          · 表头特征词命中（站号/半宽/高度/系数/水线/甲板/station/z/half 等）加权
          · 其下方必须存在含数值的数据行（排除"全文本备注行"误判）
        无可靠候选时按序回退：
          ① 首行数据比已消费表头更宽且以文本为主 → 判定为标题行下的真实表头
          ② 沿用传入的 headers（首行）
          ③ 生成默认列名 UnnamedColumn_n
        """
        self.log('开始智能识别表头行...')
        if not rows:
            return headers, []
        n_cols = max(len(headers), max((len(r) for r in rows), default=0))
        scan = min(20, len(rows))

        def _nonempty(r):
            return [v for v in r[:n_cols] if v is not None and str(v).strip() != '']

        def _texts(r):
            return [v for v in r[:n_cols] if isinstance(v, str) and v.strip()]

        def _score(r):
            ne = _nonempty(r)
            if len(ne) < 2:
                return -1.0
            ratio = len(ne) / max(n_cols, 1)
            text_ratio = len(_texts(r)) / len(ne)
            kw = ('站号', '半宽', '高度', '系数', '矩臂', '水线', '甲板', '吃水',
                  'station', 'z_long', 'half', 'deck', 'wl', '列')
            hits = sum(1 for v in ne if any(k in str(v).lower() for k in kw))
            return ratio * 2.0 + text_ratio + 1.5 * hits

        def _finalize(hdr_row):
            new_headers = []
            for i in range(n_cols):
                h = hdr_row[i] if i < len(hdr_row) else ''
                if h is None or str(h).strip() == '':
                    new_headers.append('UnnamedColumn_%d' % (i + 1))
                else:
                    new_headers.append(self._hdr_str(h).strip())
            return core.make_unique_strings(new_headers)

        def _drop_units_row(data_rows):
            """表头正下方的"单位行"（全文本且不含表头词，其下有数值行）应丢弃"""
            if not data_rows:
                return data_rows
            first = data_rows[0]
            ne = [v for v in first[:n_cols] if v is not None and str(v).strip() != '']
            if not ne:
                return data_rows
            if any(isinstance(v, (int, float, np.number)) for v in ne):
                return data_rows
            texts = [str(v).strip() for v in ne]
            kw = ('站号', '半宽', '高度', '系数', '矩臂', '水线', '甲板', '吃水',
                  'station', 'z_long', 'half', 'deck', 'wl', '列')
            if any(any(k in t.lower() for k in kw) for t in texts):
                return data_rows
            below = data_rows[1:]
            if any(any(isinstance(v, (int, float, np.number)) for v in
                       [x for x in k[:n_cols] if x is not None and str(x).strip() != ''])
                   for k in below):
                self.log('检测到表头下方的单位行，已自动跳过：%s' % ' '.join(texts))
                return below
            return data_rows

        header_row = None
        best_score = 0.0
        for r in range(scan):
            s = _score(rows[r])
            if s > best_score and s >= 5.0:
                below = rows[r + 1:]
                below_has_number = any(
                    any(isinstance(v, (int, float, np.number)) for v in _nonempty(k))
                    for k in below)
                if below_has_number or not below:
                    best_score = s
                    header_row = r

        if header_row is not None:
            self.log('在第 %d 行找到表头（评分 %.2f）。' % (header_row + 1, best_score))
            new_headers = _finalize(rows[header_row])
            data = [list(r) + [''] * (n_cols - len(r)) for r in rows[header_row + 1:]]
        else:
            # 回退 ①：标题行情形——首行数据比已消费表头更宽且以文本为主
            if rows and len(_texts(rows[0])) > len(_texts(headers)):
                self.log('首行数据比首行标题更宽且为文本，判定为表头行（标题行情形）。')
                new_headers = _finalize(rows[0])
                data = [list(r) + [''] * (n_cols - len(r)) for r in rows[1:]]
            # 回退 ②：沿用首行作为列名
            elif any(h is not None and str(h).strip() for h in headers):
                self.log('未找到更合适的表头行，沿用首行作为列名。')
                new_headers = _finalize(headers)
                data = [list(r) + [''] * (n_cols - len(r)) for r in rows]
            # 回退 ③：默认列名
            else:
                self.log('未找到合适的表头行，将使用默认列名。')
                new_headers = ['UnnamedColumn_%d' % (i + 1) for i in range(n_cols)]
                data = [list(r) + [''] * (n_cols - len(r)) for r in rows]
        data = _drop_units_row(data)
        self.log('列名已提取并唯一化：%s' % ', '.join(new_headers))
        return new_headers, data

    # =====================================================================
    # 机器学习
    # =====================================================================
    def add_ml_model_clicked(self):
        self.clear_logs()
        path = filedialog.askopenfilename(
            title='请选择 sklearn 模型文件',
            filetypes=[('Pickle 模型', '*.pkl *.pickle'), ('所有文件', '*.*')],
            parent=self.root)
        if not path:
            self.log('加载模型操作已取消。')
            return
        if not path.lower().endswith(('.pkl', '.pickle')):
            messagebox.showerror(
                '不支持的格式',
                '本工程已不再支持 MATLAB 导出的 .mat 模型。\n\n'
                '请使用工具栏【训练模型】或右键【训练模型...】\n'
                '基于已标注的型值表用 scikit-learn 训练模型，\n'
                '并保存为 .pkl 后再加载。',
                parent=self.root)
            return
        try:
            self.log('正在加载 sklearn 模型: %s' % path)
            model, req_vars = ml_utils.load_model(path)
            self.ML_model = {'model': model, 'required_variables': list(req_vars),
                             'kind': 'Python', 'path': path}
            self.log('模型加载成功！特征维度: %d' % len(req_vars))
            # 创建 Model 节点
            node = self._tree_add(self.model_root, os.path.basename(path),
                                  {'type': 'model', 'path': path})
            self.log('模型信息已更新到左侧树状图中。')
            self._refresh_statusbar()
            self._update_button_state()
            self._mark_dirty()
            messagebox.showinfo('操作成功', '模型加载成功！', parent=self.root)
        except Exception as e:
            messagebox.showerror('加载失败', '加载模型失败: %s' % e, parent=self.root)

    def extract_clicked(self):
        """特征提取 + 列角色预测（MATLAB PushTool_extractClicked）"""
        self.clear_logs()
        if self.ML_model is None:
            messagebox.showerror('操作失败', '请先加载一个模型。', parent=self.root)
            return
        sel = self._selected_node()
        meta = self.tree_meta.get(sel) if sel else None
        if not meta or 'Data' not in meta:
            messagebox.showerror('操作无效', '请先选择一个有效的表格节点。', parent=self.root)
            return
        self.log('开始从表格 "%s" 提取数据...' % self.tree.item(sel, 'text'))
        original_data = meta['Data']
        var_names = meta.get('VariableNames', meta.get('Headers', []))
        req_vars = self.ML_model.get('required_variables', ml_utils.DEFAULT_FEATURES)

        # 数值化
        numeric = np.array([[core._parse_scalar(c) for c in row] for row in original_data],
                           dtype=float)
        n_cols = numeric.shape[1]
        feats, warnings = ml_utils.extract_prediction_features(numeric, req_vars)
        for w in warnings:
            self.log(w)
        # 防御性填充：确保喂给模型的任何列都不含 NaN（与训练口径一致）
        from src.core.classifier import ShipColumnClassifier
        ShipColumnClassifier._fill_feature_nan(feats)
        model = self.ML_model.get('model')
        if model is None:
            messagebox.showerror('无法预测', '该模型没有可用的预测后端。\n'
                                '请在"模型"节点右键选择"训练模型"，使用 scikit-learn '
                                '重新训练后再执行特征提取。', parent=self.root)
            return
        labels = np.asarray(model.predict(feats), dtype=object).ravel()
        self.log('预测结果: %s' % list(labels))
        target = {'station', 'z', 'half'}
        keep_idx = [i for i, lb in enumerate(labels) if str(lb).lower() in target]

        if not keep_idx:
            messagebox.showerror('提取失败', '模型未能识别出任何 station/z/half 列。',
                                 parent=self.root)
            return
        initial_roles = [str(labels[i]).lower() for i in keep_idx]
        initial_data = [[original_data[r][i] for i in keep_idx] for r in range(len(original_data))]
        # 用户修正角色
        names = [var_names[i] if i < len(var_names) else '列_%d' % (i + 1) for i in keep_idx]
        roles = self._ask_roles_dialog(names, initial_roles)
        if roles is None:
            self.log('操作已取消。')
            return
        is_kept = [r not in ('NULL', 'none', '') for r in roles]
        final_roles = [roles[i] for i in range(len(roles)) if is_kept[i]]
        final_data = [[row[i] for i in range(len(keep_idx)) if is_kept[i]]
                      for row in initial_data]
        final_idx = [keep_idx[i] for i in range(len(keep_idx)) if is_kept[i]]
        final_numeric = numeric[:, final_idx]

        s_c = hw_c = hd_c = zd_c = zl_c = 0
        base_names = []
        user_names = []
        for r in final_roles:
            rl = r.lower()
            if rl == 'station':
                s_c += 1
                base_names.append('station_%d' % s_c)
                user_names.append('站号列_%d' % s_c)
            elif rl == 'half_wl':
                hw_c += 1
                base_names.append('half_wl_%d' % hw_c)
                user_names.append('水线半宽_%d' % hw_c)
            elif rl == 'half_deck':
                hd_c += 1
                base_names.append('half_deck_%d' % hd_c)
                user_names.append('甲板半宽_%d' % hd_c)
            elif rl == 'z_deck':
                zd_c += 1
                base_names.append('z_deck_%d' % zd_c)
                user_names.append('甲板高度_%d' % zd_c)
            elif rl == 'z_long':
                zl_c += 1
                base_names.append('z_long_%d' % zl_c)
                user_names.append('纵剖线坐标_%d' % zl_c)
            else:
                base_names.append('col_%d' % len(base_names))
                user_names.append('列_%d' % len(user_names))
        self.log('正在配置水线固定高度...')
        height_values = np.full(len(final_roles), np.nan)
        wl_indices = [i for i, r in enumerate(final_roles) if r.lower() == 'half_wl']
        # 严格校验 1：必须识别到水线半宽列，否则后续水线面/浮心/三维计算无法进行
        if not wl_indices:
            if not messagebox.askyesno(
                    '未识别到水线半宽列',
                    '本次识别结果中没有任何 half_wl（水线半宽）列。\n\n'
                    '缺少水线高度将无法进行：\n'
                    '  · 水线面计算 / 静水力曲线\n'
                    '  · 浮心计算（水线面法）\n'
                    '  · 三维型线与船体蒙皮\n\n'
                    '仍要按当前角色生成"识别结果"吗？\n'
                    '（选"否"返回重新修正列角色）', parent=self.root):
                self.log('操作已取消：未识别到水线半宽列。')
                return
        # 严格校验 2：逐条水线询问高度，非法输入必须重输，取消则中止
        used_heights = []
        default_h = 0.0
        for i in wl_indices:
            v = self._ask_waterline_height(user_names[i], default_h, used_heights)
            if v is None:
                self.log('操作已取消：未完成水线高度输入。')
                return
            height_values[i] = v
            used_heights.append(v)
            self.log('水线 %s 固定高度 = %g m' % (user_names[i], v))
            default_h = v + 1.0
        # 创建结果节点
        parent_node = self._tree_add(self.table_root, '提取自 - %s' % self.tree.item(sel, 'text'),
                                     {'type': 'extraction'})
        result_node = self._tree_add(parent_node, '识别结果', {
            'type': 'result', 'Roles': final_roles, 'BaseNames': base_names,
            'UserNames': user_names, 'Data': final_data, 'Numeric': final_numeric,
            'heightValues': height_values})
        self.original_headers = user_names
        self.original_data = final_data
        self.original_table.set_columns(user_names)
        self.original_table.set_data(final_data)
        self.tree.selection_set(result_node)
        messagebox.showinfo('操作成功', '数据提取和角色配置成功！', parent=self.root)

    def _ask_waterline_height(self, name, default=0.0, used_heights=None):
        """严格询问某条水线(WL)的固定高度。

        - 必须是有限非负数（距基线，单位 m），非法输入报错并要求重输
        - 不得与已输入的水线高度重复（同高水线会导致几何退化）
        - 取消/留空 → 二次确认；确认取消则返回 None（调用方应中止整个流程）
        """
        used_heights = used_heights if used_heights is not None else []
        while True:
            ans = ask_text_dialog(
                self.root, '设置水平水线(WL)的固定高度',
                '请输入"%s"的固定高度 (m，距基线，非负数):' % name,
                '%g' % default)
            if ans is None or not str(ans).strip():
                if messagebox.askyesno(
                        '缺少水线高度',
                        '水线高度是后续"水线面计算 / 静水力曲线 / 浮心（水线面法）/\n'
                        '三维型线与船体蒙皮"的必需输入，缺少将无法进行这些计算。\n\n'
                        '确定要中止本次列识别吗？\n'
                        '（选"否"返回继续输入高度）', parent=self.root):
                    return None
                continue
            try:
                v = float(str(ans).strip())
            except ValueError:
                messagebox.showerror('输入无效',
                                     '"%s" 不是有效数字，请重新输入。' % str(ans).strip(),
                                     parent=self.root)
                continue
            if not math.isfinite(v) or v < 0:
                messagebox.showerror(
                    '输入无效',
                    '水线高度必须为有限的非负数（距基线，单位 m）。\n'
                    '当前输入：%g' % v, parent=self.root)
                continue
            if any(abs(v - u) < 1e-9 for u in used_heights):
                messagebox.showerror(
                    '水线高度重复',
                    '已有水线使用了高度 %g m。\n'
                    '同一高度的多条水线会导致几何退化，请重新输入。' % v,
                    parent=self.root)
                continue
            return v

    def edit_waterline_heights_clicked(self):
        """在"识别结果"节点上重新输入各水线的固定高度（无需重新识别）"""
        sel = self._selected_node()
        meta = self.tree_meta.get(sel) if sel else None
        if not meta or meta.get('type') != 'result':
            messagebox.showinfo('提示', '请先选择"识别结果"节点。', parent=self.root)
            return
        roles = meta.get('Roles', [])
        user_names = meta.get('UserNames', [])
        heights = meta.get('heightValues', None)
        if heights is None:
            heights = np.full(len(roles), np.nan)
        wl_indices = [i for i, r in enumerate(roles) if str(r).lower() == 'half_wl']
        if not wl_indices:
            messagebox.showinfo('提示',
                                '该识别结果不包含 half_wl（水线半宽）列，无需设置水线高度。',
                                parent=self.root)
            return
        used = [float(heights[i]) for i in wl_indices
                if i < len(heights) and math.isfinite(heights[i])]
        for i in wl_indices:
            cur = float(heights[i]) if i < len(heights) and math.isfinite(heights[i]) else 0.0
            others = [u for u in used if abs(u - cur) > 1e-9] if math.isfinite(cur) else used
            v = self._ask_waterline_height(user_names[i], cur, others)
            if v is None:
                self.log('水线高度编辑已取消。')
                return
            if math.isfinite(cur):
                used = [u for u in used if abs(u - cur) > 1e-9]
            heights[i] = v
            used.append(v)
            self.log('水线 %s 固定高度 = %g m' % (user_names[i], v))
        meta['heightValues'] = heights
        self.tree_meta[sel] = meta
        self._mark_dirty()
        self.log('水线高度已更新，可继续"从offset导入"与后续计算。')
        messagebox.showinfo('完成', '水线高度已更新。', parent=self.root)

    def _ask_roles_dialog(self, names, initial_roles):
        """角色修正对话框，返回最终角色列表或 None"""
        dlg = tk.Toplevel(self.root)
        dlg.title('修正列属性')
        dlg.transient(self.root)
        dlg.grab_set()
        result = {'value': None}
        tk.Label(dlg, text='为每一列指定角色：').pack(anchor='w', padx=8, pady=(8, 2))
        role_options = ['station', 'half_wl', 'half_deck', 'z_deck', 'z_long', 'NULL']
        rows_frame = ttk.Frame(dlg)
        rows_frame.pack(fill='both', expand=True, padx=8, pady=4)
        vars_ = []
        for i, name in enumerate(names):
            tk.Label(rows_frame, text=str(name), width=18, anchor='w').grid(row=i, column=0, pady=1)
            v = tk.StringVar(value=initial_roles[i] if i < len(initial_roles) else 'NULL')
            ttk.Combobox(rows_frame, textvariable=v, values=role_options, width=12,
                         state='readonly').grid(row=i, column=1, pady=1, padx=6)
            vars_.append(v)

        def ok():
            result['value'] = [v.get() for v in vars_]
            dlg.destroy()

        btns = ttk.Frame(dlg, padding=8)
        btns.pack()
        ttk.Button(btns, text='确定', command=ok).pack(side='left', padx=6)
        ttk.Button(btns, text='取消', command=dlg.destroy).pack(side='left', padx=6)
        self.root.wait_window(dlg)
        return result['value']

    def _collect_data_nodes(self):
        """
        收集所有含表格数据（Data）的树节点，用于训练数据选择。
        返回 list of (iid, meta, display_text)
        """
        result = []

        def walk(node):
            for child in self.tree.get_children(node):
                meta = self.tree_meta.get(child, {})
                if 'Data' in meta and meta['Data']:
                    text = self.tree.item(child, 'text')
                    ntype = meta.get('type', '')
                    kind = {'table': '表格', 'result': '识别结果'}.get(ntype, '数据')
                    result.append((child, meta, '%s [%s]' % (text, kind)))
                walk(child)

        walk(self.tree_root)
        return result

    @staticmethod
    def _detect_label_row(data):
        """
        自动检测包含角色标签的行（支持中文别名与大小写不敏感）。
        data: list[list]   表格数据（不含表头）
        返回: (行索引 0基, 识别到的标签 dict {col_idx: normalized_label})
              或 (None, {}) 表示未检测到
        """
        from src.core.classifier import normalize_label
        best_idx = None
        best_count = 0
        for i, row in enumerate(data):
            count = 0
            for v in row:
                if normalize_label(v) is not None:
                    count += 1
            if count > best_count:
                best_count = count
                best_idx = i
        if best_idx is None:
            return None, {}
        labels = {j: normalize_label(v) for j, v in enumerate(data[best_idx])}
        return best_idx, labels

    def train_model_clicked(self):
        """
        训练列角色分类模型（KNN/SVM/Tree）。

        修复要点：
        1. 训练数据选择：优先使用当前选中节点；若无数据则从所有含数据的节点中
           弹窗选择（原先只报错"请先选择表格节点"，而菜单挂在 Model 节点下导致无法选中）。
        2. 特征一致性：每个【列】作为一个样本，提取 12 个统计特征，与预测时
           extract_prediction_features 的输出完全一致（原先用每行原始值，维度不匹配）。
        """
        # ---- 1. 获取候选数据节点 ----
        candidates = self._collect_data_nodes()
        if not candidates:
            messagebox.showerror(
                '无数据',
                '未找到可用于训练的表格数据。\n\n'
                '请先点击工具栏【导入表格】导入型值表，\n'
                '表中每一列代表一个已知角色的样本列，并包含一行角色标签\n'
                '（如 station / z / half）。',
                parent=self.root)
            return

        # ---- 2. 选择训练数据节点 ----
        sel = self._selected_node()
        meta = self.tree_meta.get(sel) if sel else None
        if not meta or not meta.get('Data'):
            # 当前选中节点无数据 → 弹窗让用户选择
            names = [t for _, _, t in candidates]
            idx = ask_choice_dialog(self.root, '选择训练数据',
                                    '请选择作为训练数据的表格节点：', names)
            if idx is None:
                return
            sel, meta, _ = candidates[idx]
            self.tree.selection_set(sel)

        data = meta['Data']
        headers = meta.get('VariableNames', meta.get('Headers', []))
        if not data:
            messagebox.showerror('无数据', '所选节点不含数据行。', parent=self.root)
            return

        # ---- 3. 确定标签行（自动检测 + 可选手动指定） ----
        label_row, label_dict = self._detect_label_row(data)
        if label_row is None:
            # 未自动检测到 → 列出前若干行供用户选择
            preview = []
            max_show = min(10, len(data))
            for i in range(max_show):
                row_text = ' | '.join(str(v)[:14] for v in data[i][:6])
                preview.append('第 %d 行: %s' % (i + 1, row_text))
            idx = ask_choice_dialog(
                self.root, '选择角色标签行',
                '未自动检测到角色标签行（station/z/half 或 站号/高度/半宽）。\n'
                '请选择包含角色标签的行：', preview)
            if idx is None:
                return
            label_row = idx
            from src.core.classifier import normalize_label
            label_dict = {j: normalize_label(v) for j, v in enumerate(data[label_row])}

        # ---- 4. 提取每列的 12 个统计特征（与预测一致） ----
        # 保留有合法标签的列
        keep_cols = [j for j, lb in label_dict.items() if lb is not None]
        valid_labels = [label_dict[j] for j in keep_cols]

        # 诊断：每列的标签
        all_labels = [(j, data[label_row][j] if j < len(data[label_row]) else '',
                       label_dict.get(j)) for j in range(len(headers))]
        n_invalid = len(headers) - len(keep_cols)
        invalid_summary = ''
        if n_invalid > 0:
            samples = [f"列 {i}: '{raw}'"
                       for i, raw, lb in all_labels
                       if lb is None][:5]
            invalid_summary = '\n未识别列示例：\n' + '\n'.join(samples)
        if len(keep_cols) < 2:
            messagebox.showerror(
                '数据不足',
                '有效标注列不足（需 ≥ 2 列）。\n'
                '识别到 %d 列有效，%d 列无效。\n\n'
                '有效标签：station / z / half\n'
                '（或别名 站号 / 高度 / 半宽），不区分大小写。'
                + invalid_summary,
                parent=self.root)
            return

        # 数据行 = 除标签行外的所有行
        data_rows = [row for i, row in enumerate(data) if i != label_row]
        if len(data_rows) < 3:
            messagebox.showerror('数据不足', '有效数据行不足（需 ≥ 3 行）', parent=self.root)
            return

        # 构造数值矩阵 (n_rows, n_cols)，仅保留有效列
        matrix = np.full((len(data_rows), len(keep_cols)), np.nan)
        for i, row in enumerate(data_rows):
            for k, j in enumerate(keep_cols):
                v = row[j] if j < len(row) else None
                if v is None or v == '':
                    continue
                try:
                    matrix[i, k] = float(v)
                except (ValueError, TypeError):
                    pass

        from src.core.classifier import ShipColumnClassifier
        clf = ShipColumnClassifier()
        clf.required_variables = list(ml_utils.DEFAULT_FEATURES)
        feats, warnings = clf.extract_features(matrix)  # (n_cols, 12)
        for w in warnings:
            self.log('  特征警告: %s' % w)
        y = np.array(valid_labels, dtype=object)

        # 移除含 NaN 特征的样本
        valid_samples = ~np.any(np.isnan(feats), axis=1)
        if np.sum(valid_samples) < 2:
            messagebox.showerror('数据不足',
                                 '有效样本不足（需 ≥ 2 列），无法训练。',
                                 parent=self.root)
            return
        feats = feats[valid_samples]
        y = y[valid_samples]

        # ---- 5. 选择模型类型 ----
        model_type = ask_choice_dialog(
            self.root, '选择模型类型',
            '请选择要训练的分类器类型：',
            ['KNN（K近邻）', 'SVM（支持向量机）', 'Tree（决策树）'], default_index=0)
        if model_type is None:
            return
        type_map = {0: 'KNN', 1: 'SVM', 2: 'Tree'}
        model_type = type_map[model_type]

        # ---- 6. 训练 ----
        try:
            model = ml_utils.train_model(feats, y, model_type)
        except Exception as e:
            messagebox.showerror('训练失败', '训练失败: %s' % e, parent=self.root)
            return

        # 与预测端保持一致的 12 维特征
        self.ML_model = {
            'model': model,
            'required_variables': list(ml_utils.DEFAULT_FEATURES),
            'kind': 'Python',
            'path': None,
            'feature_dim': feats.shape[1],
            'train_samples': int(feats.shape[0]),
        }
        self.log('模型训练成功！类型: %s, 样本(列)数: %d, 特征数: %d'
                 % (model_type.upper(), feats.shape[0], feats.shape[1]))
        self._tree_add(self.model_root, '训练模型(%s)' % model_type.upper(),
                       {'type': 'model', 'path': None, 'kind': 'Python'})
        self._refresh_statusbar()
        self._update_button_state()
        msg = '模型训练完成！\n\n类型: %s\n样本(列)数: %d\n特征数: %d\n\n' \
              '现在可选中表格节点后点击工具栏【特征提取】进行列角色预测。' \
              % (model_type.upper(), int(feats.shape[0]), int(feats.shape[1]))
        self._mark_dirty()
        messagebox.showinfo('成功', msg, parent=self.root)

    # =====================================================================
    # 树操作
    # =====================================================================
    def _selected_node(self):
        sel = self.tree.selection()
        return sel[0] if sel else None

    def tree_selection_changed(self, event):
        sel = self._selected_node()
        if not sel:
            self.original_table.set_data([])
            self._update_button_state()
            self._refresh_statusbar()
            return
        # 选中项目或其分组时切换到该项目（新数据挂到该项目下）
        proj = self._node_project(sel)
        if proj and proj != self._current_project:
            self._switch_project(proj)
        meta = self.tree_meta.get(sel, {})
        ntype = meta.get('type', '')
        if ntype == 'waterline':
            # 切换到半宽表格显示水线面数据
            table = meta.get('table')
            if table:
                self.notebook.select(1)
                self.Half_table.set_columns(table.get('columns', ['列', '站号', '半宽', '系数', '相对矩臂']))
                self.Half_table.set_data(table.get('rows', []))
                self.update_half_width_plot()
        elif ntype == 'bodyplan':
            table = meta.get('table')
            if table:
                self.notebook.select(2)
                self.Z_table.set_columns(table.get('columns', ['列', '高度', '半宽', '系数']))
                self.Z_table.set_data(table.get('rows', []))
                self.update_transverse_section_plot()
        elif ntype == 'result':
            # 与 MATLAB Tree_tableSelectionChanged 一致：选中"识别结果"节点时只
            # 刷新原表格内容，不切换 Tab。
            # 若在此处切到"原表格"页，isTap 会被重置为 1，
            # 导致"从offset导入"的 isTap_2(水线面)/isTap_3(横剖面) 分支永远无法命中。
            data = meta.get('Data', [])
            names = meta.get('UserNames', meta.get('BaseNames', []))
            if names:
                self.original_headers = names
                self.original_data = data
                self.original_table.set_columns(names)
                self.original_table.set_data(data)
        elif ntype == 'table':
            data = meta.get('Data', [])
            names = meta.get('Headers', [])
            self.original_headers = names
            self.original_data = data
            self.original_table.set_columns(names)
            self.original_table.set_data(data)
        # 刷新按钮启用与状态栏
        self._update_button_state()
        self._refresh_statusbar()

    def tree_double_click(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            self._rename_tree_node(iid)

    def tree_right_click(self, event):
        iid = self.tree.identify_row(event.y)
        menu = tk.Menu(self.root, tearoff=0)
        if not iid:
            # 空白处：新建项目
            menu.add_command(label='新建项目', command=self._new_project)
            menu.tk_popup(event.x_root, event.y_root)
            return
        self.tree.selection_set(iid)
        meta = self.tree_meta.get(iid, {})
        ntype = meta.get('type', '')

        # 所有节点均可重命名（含项目/分组）
        menu.add_command(label='重命名',
                         command=lambda: self._rename_tree_node(iid))

        # 分组（Table/Model/Face）不可删除；项目可删除（至少保留一个）
        is_group = ntype in ('table_root', 'model_root', 'face_root')
        if not is_group:
            menu.add_command(label='删除节点',
                             command=lambda: self._delete_tree_node(iid))

        # 训练模型：Model / Table / result 节点均可发起
        # （内部会自动选择训练数据，无需用户先"选中表格"）
        if ntype in ('model', 'table', 'result') or iid == self.model_root:
            menu.add_command(label='训练模型...', command=self.train_model_clicked)

        # 含数据的节点：可直接作为训练数据
        if meta.get('Data'):
            menu.add_command(label='用此表训练模型...',
                             command=self.train_model_clicked)

        if ntype == 'result':
            menu.add_command(label='从offset导入...',
                             command=self.import_from_offset_clicked)
            menu.add_command(label='编辑水线高度...',
                             command=self.edit_waterline_heights_clicked)
            menu.add_command(label='编辑列名...', command=self.edit_colnames_clicked)

        if ntype == 'table':
            menu.add_command(label='特征提取（预测列角色）...',
                             command=self.extract_clicked)

        menu.tk_popup(event.x_root, event.y_root)

    def _delete_tree_node(self, iid):
        # 项目节点：可删除（至少保留一个），若删除当前项目则切换到剩余项目
        if self.tree_meta.get(iid, {}).get('type') == 'project':
            projects = [c for c in self.tree.get_children('')
                        if self.tree_meta.get(c, {}).get('type') == 'project']
            if len(projects) <= 1:
                messagebox.showinfo('操作无效', '至少保留一个项目。', parent=self.root)
                return
            if iid == self._current_project:
                others = [c for c in projects if c != iid]
                self._switch_project(others[0])
            self.tree.delete(iid)
            self.tree_meta.pop(iid, None)
            self._update_button_state()
            self._refresh_statusbar()
            self.log('项目已删除。')
            return
        # 分组节点不可删除
        if self.tree_meta.get(iid, {}).get('type') in ('table_root', 'model_root', 'face_root'):
            messagebox.showinfo('操作无效', '不能删除分组节点。', parent=self.root)
            return
        self.tree.delete(iid)
        self.tree_meta.pop(iid, None)
        self._update_button_state()
        self._refresh_statusbar()
        self.log('节点已删除。')

    def edit_colnames_clicked(self):
        sel = self._selected_node()
        meta = self.tree_meta.get(sel) if sel else None
        if not meta or meta.get('type') != 'result':
            messagebox.showinfo('提示', '请先选择"识别结果"节点。', parent=self.root)
            return
        names = meta.get('UserNames', meta.get('BaseNames', []))
        from src.ui.ui_widgets import ask_text_dialog
        for i in range(len(names)):
            new = ask_text_dialog(self.root, '编辑列名', '列 %d 当前名 "%s" 的新名称:' % (i + 1, names[i]), names[i])
            if new is None:
                return
            names[i] = new
        meta['UserNames'] = names
        self.tree_meta[sel] = meta
        self.original_headers = names
        self.original_table.set_columns(names)
        self._mark_dirty()
        self.log('列名已更新。')

    def import_from_offset_clicked(self):
        """从识别结果导入型线

        与 MATLAB PushTool_import_from_offsetClicked 一致，按当前所在 Tab 决定导入目标：
        - 半宽页   (isTap=2, isTap_2) → 水线面 / 甲板线
        - 横剖面页 (isTap=3, isTap_3) → 横剖面 (Body Plan)
        """
        sel = self._selected_node()
        meta = self.tree_meta.get(sel) if sel else None
        if not meta or meta.get('type') != 'result':
            messagebox.showerror('操作无效', '请先选择一个"识别结果"节点。', parent=self.root)
            return

        if self.isTap == 2:
            self._import_waterlines_from_result(meta)
        elif self.isTap == 3:
            self.import_bodyplans_clicked(meta)
        else:
            messagebox.showinfo(
                '导入型线',
                '请切换到"半宽"页导入水线面/甲板线，\n'
                '或切换到"横剖面"页导入横剖面。', parent=self.root)

    def _find_or_create_child(self, parent, text, meta=None):
        """在 parent 下查找同名子节点并复用，不存在才新建

        对应 MATLAB 的 findobj(...Children,'Text',...) + isempty(...) 建节点模式，
        避免重复导入时产生多个"船型模型 / Half / Body Plan"节点。
        """
        for child in self.tree.get_children(parent):
            if self.tree.item(child, 'text') == text:
                return child
        return self._tree_add(parent, text, meta)

    def _import_waterlines_from_result(self, meta):
        """从识别结果导入水线面 / 甲板线（半宽页）"""
        roles = meta.get('Roles', [])
        user_names = meta.get('UserNames', [])
        data = meta.get('Data', [])
        numeric = meta.get('Numeric', None)
        heights = meta.get('heightValues', [])
        station_idx = None
        half_wl = []
        half_deck = []
        z_deck = []
        for i, r in enumerate(roles):
            rl = str(r).lower()
            if rl == 'station':
                station_idx = i
            elif rl == 'half_wl':
                half_wl.append(i)
            elif rl == 'half_deck':
                half_deck.append(i)
            elif rl == 'z_deck':
                z_deck.append(i)
        if station_idx is None:
            messagebox.showerror('导入失败', '源数据缺少 station 列。', parent=self.root)
            return
        station_data = np.array([core._parse_scalar(r[station_idx]) for r in data], dtype=float)
        if numeric is None:
            numeric = np.array([[core._parse_scalar(c) for c in row] for row in data], dtype=float)
        # 水线面 + 甲板线
        items = []
        item_types = []
        for i in half_wl:
            items.append('%s (水线)' % user_names[i])
            item_types.append('wl')
        for i in half_deck:
            items.append('%s (甲板)' % user_names[i])
            item_types.append('deck')
        if not items:
            messagebox.showerror('导入失败', '源数据缺少 half_wl/half_deck 列。', parent=self.root)
            return
        # 严格校验：拟导入的水线必须有有效的固定高度，否则后续计算无法进行
        sel_idx = ask_multi_select(self.root, '选择导入项', '请选择要导入的线型：', ['-- 全部导入 --'] + items)
        if sel_idx is None:
            return
        if 1 in sel_idx:
            final_idx = list(range(len(items)))
        else:
            final_idx = [i - 1 for i in sel_idx]
        wl_overall = [sel_overall for k in final_idx
                      for sel_overall in [(half_wl + half_deck)[k]] if item_types[k] == 'wl']
        bad_wl = [user_names[i] for i in wl_overall
                  if not (i < len(heights) and math.isfinite(heights[i]) and heights[i] >= 0)]
        if bad_wl:
            messagebox.showerror(
                '缺少水线高度',
                '以下水线未设置有效的固定高度，无法导入：\n  %s\n\n'
                '请在"识别结果"节点右键选择【编辑水线高度...】补齐后重试。'
                % '\n  '.join(bad_wl), parent=self.root)
            return
        # 创建 Face -> 船型模型 -> Half 节点（已存在则复用）
        ship_node = self._find_or_create_child(
            self.face_root, '船型模型', {'type': 'ship_model_root'})
        half_node = self._find_or_create_child(ship_node, 'Half', {'type': 'half_root'})
        success = 0
        for k in final_idx:
            sel_overall = (half_wl + half_deck)[k]
            sel_type = item_types[k]
            col_name = user_names[sel_overall]
            half_data = numeric[:, sel_overall] / 1000.0
            if sel_type == 'wl':
                valid = np.isfinite(station_data) & np.isfinite(half_data)
                st = station_data[valid]
                hw = half_data[valid]
                if st.size == 0:
                    continue
                rows = [[i + 1, core.num2trimstr(st[i]), core.num2trimstr(hw[i]), '', '']
                        for i in range(st.size)]
                table = {'columns': ['列', '站号', '半宽', '系数', '相对矩臂'], 'rows': rows}
                h = heights[sel_overall] if sel_overall < len(heights) and math.isfinite(
                    heights[sel_overall]) else math.nan
                self.waterlines.append({'type': 'waterline', 'name': col_name, 'height': h, 'table': table})
                self._tree_add(half_node, col_name, {'type': 'waterline', 'table': table,
                                                     'height': h})
            else:
                deck_rank = half_deck.index(sel_overall)
                if deck_rank >= len(z_deck):
                    continue
                z_data = numeric[:, z_deck[deck_rank]] / 1000.0
                valid = np.isfinite(station_data) & np.isfinite(half_data) & np.isfinite(z_data)
                st = station_data[valid]
                hw = half_data[valid]
                zd = z_data[valid]
                if st.size == 0:
                    continue
                rows = [[i + 1, core.num2trimstr(st[i]), core.num2trimstr(hw[i]),
                         core.num2trimstr(zd[i])] for i in range(st.size)]
                table = {'columns': ['列', '站号', '半宽', '高度'], 'rows': rows}
                self.decklines.append({'type': 'deckline', 'name': col_name, 'table': table})
                self._tree_add(half_node, col_name, {'type': 'deckline', 'table': table})
            success += 1
        self.tree.item(self.face_root, open=True)
        self.tree.item(ship_node, open=True)
        self.tree.item(half_node, open=True)
        self.update_half_width_plot()
        self._update_button_state()
        self._mark_dirty()
        messagebox.showinfo('导入完成', '成功导入 %d 条曲线。' % success, parent=self.root)

    def import_bodyplans_clicked(self, meta, numeric=None, station_data=None):
        """从识别结果导入横剖面到 Body Plan 节点（横剖面页）"""
        roles = meta.get('Roles', [])
        heights = meta.get('heightValues', [])
        data = meta.get('Data', [])
        half_wl = [i for i, r in enumerate(roles) if str(r).lower() == 'half_wl']
        half_deck = [i for i, r in enumerate(roles) if str(r).lower() == 'half_deck']
        z_deck = [i for i, r in enumerate(roles) if str(r).lower() == 'z_deck']
        station_idx = next((i for i, r in enumerate(roles)
                            if str(r).lower() == 'station'), None)
        if station_idx is None:
            messagebox.showerror('导入失败', '源数据缺少 station 列。', parent=self.root)
            return
        if numeric is None:
            numeric = meta.get('Numeric', None)
        if numeric is None:
            numeric = np.array([[core._parse_scalar(c) for c in row] for row in data], dtype=float)
        if station_data is None:
            station_data = np.array(
                [core._parse_scalar(r[station_idx]) for r in data], dtype=float)
        valid_station_rows = np.nonzero(np.isfinite(station_data))[0]
        if valid_station_rows.size == 0:
            messagebox.showerror('无数据', '"station"列中无有效数据。', parent=self.root)
            return
        items = [core.num2trimstr(station_data[i]) for i in valid_station_rows]
        sel_idx = ask_multi_select(self.root, '选择站', '请选择要导入的横剖面站：', ['-- 全部导入 --'] + items)
        if sel_idx is None:
            return
        if 1 in sel_idx:
            final_idx = list(range(len(items)))
        else:
            final_idx = [i - 1 for i in sel_idx]
        # Face -> 船型模型 -> Body Plan（已存在则复用，避免重复导入产生多余节点）
        ship_node = self._find_or_create_child(
            self.face_root, '船型模型', {'type': 'ship_model_root'})
        body_node = self._find_or_create_child(ship_node, 'Body Plan', {'type': 'bodyplan_root'})
        # 严格校验：若水线高度缺失且没有甲板高度列，剖面将没有任何点可导入
        wl_ok = [i for i in half_wl
                 if i < len(heights) and math.isfinite(heights[i]) and heights[i] >= 0]
        if half_wl and not wl_ok and not half_deck:
            messagebox.showerror(
                '缺少水线高度',
                '所有 half_wl（水线半宽）列均未设置有效固定高度，\n'
                '且没有甲板高度(z_deck)列，横剖面将没有任何数据点。\n\n'
                '请在"识别结果"节点右键选择【编辑水线高度...】补齐后重试。',
                parent=self.root)
            return
        success = 0
        skipped = []
        for k in final_idx:
            row_idx = int(valid_station_rows[k])
            points = []
            for idx in half_wl:
                y = numeric[row_idx, idx]
                z = heights[idx] if idx < len(heights) else math.nan
                if math.isfinite(y) and math.isfinite(z):
                    points.append([y, z])
            if len(half_deck) == len(z_deck):
                for i in range(len(half_deck)):
                    y = numeric[row_idx, half_deck[i]]
                    z = numeric[row_idx, z_deck[i]] / 1000.0
                    if math.isfinite(y) and math.isfinite(z):
                        points.append([y, z])
            if not points:
                skipped.append(items[k])
                continue
            points = np.array(points)
            order = np.argsort(points[:, 1])
            sorted_hb = points[order, 0] / 1000.0
            sorted_z = points[order, 1]
            rows = [[i + 1, core.num2trimstr(sorted_z[i]), core.num2trimstr(sorted_hb[i]), '']
                    for i in range(len(sorted_z))]
            station_num = float(station_data[row_idx])
            table = {'columns': ['列', '高度', '半宽', '系数'], 'rows': rows}
            self.bodyplans.append({'type': 'bodyplan', 'name': '站 %s' % items[k],
                                   'station': station_num, 'table': table})
            self.sections[station_num] = {'Y': sorted_hb.tolist(), 'Z': sorted_z.tolist()}
            self._tree_add(body_node, '站 %s' % items[k],
                           {'type': 'bodyplan', 'table': table, 'station': station_num})
            success += 1
        self.tree.item(self.face_root, open=True)
        self.tree.item(ship_node, open=True)
        self.tree.item(body_node, open=True)
        self.update_transverse_section_plot()
        self._update_button_state()
        if success == 0:
            messagebox.showerror(
                '导入失败',
                '未能导入任何横剖面。\n'
                '可能原因：水线/甲板高度缺失或该站半宽数据无效。\n'
                + ('已跳过：%s' % ', '.join(skipped) if skipped else ''),
                parent=self.root)
            return
        if skipped:
            self.log('已跳过 %d 个无有效数据的站：%s' % (len(skipped), ', '.join(skipped)))
        self.log('横剖面导入完成：%d 个站。' % success)
        self._mark_dirty()
        messagebox.showinfo('导入完成', '成功导入 %d 个横剖面。' % success, parent=self.root)

    # =====================================================================
    # 半宽表格操作
    # =====================================================================
    def set_principal_clicked(self):
        def f(v):
            return '' if not math.isfinite(v) else core.num2trimstr(v, 1e-9)
        vals = ask_numeric_dialog(
            self.root, '输入主尺度',
            ['垂线间长 Lpp (m):', '型宽 Breadth (m):', '型深 Depth (m):',
             '垂线间长起始站号:', '垂线间长结束站号:'],
            [f(self.Lpp), f(self.Breadth), f(self.Depth),
             f(self.LppStartStation), f(self.LppEndStation)])
        if vals is None:
            return
        lpp, b, d, s, e = vals
        if not (math.isfinite(lpp) and lpp > 0):
            messagebox.showerror('输入错误', 'Lpp 必须为正数。', parent=self.root)
            return
        if not (math.isfinite(b) and b > 0):
            messagebox.showerror('输入错误', '型宽 Breadth 必须为正数。', parent=self.root)
            return
        if not (math.isfinite(d) and d > 0):
            messagebox.showerror('输入错误', '型深 Depth 必须为正数。', parent=self.root)
            return
        if not math.isfinite(s) or not math.isfinite(e):
            messagebox.showerror('输入错误', '站号必须为有效数值。', parent=self.root)
            return
        if e < s:
            messagebox.showerror('输入错误', '结束站号不能小于起始站号。', parent=self.root)
            return
        self.Lpp = lpp
        self.Breadth = b
        self.Depth = d
        self.LppStartStation = s
        self.LppEndStation = e
        self.log('主尺度 Lpp, Breadth, Depth 已更新。')
        self.log('  Lpp=%.3f m, B=%.3f m, D=%.3f m, 站号 %.1f~%.1f' % (lpp, b, d, s, e))
        self.update_half_width_plot()
        self._refresh_statusbar()
        self._mark_dirty()
        messagebox.showinfo('成功', '主尺度设置成功。', parent=self.root)

    def add_station_num_clicked(self):
        """快速站号（支持不等距区间 + 显式站号）"""
        from src.ui.ui_widgets import ask_text_dialog
        input_str = ask_text_dialog(self.root, '快速站号（不等距支持）',
                                    '输入区间段(起点:终点:步长)，多段用分号分隔；'
                                    '或显式站号(逗号分隔)：\n示例: 0:10:1; 或 0,1,2,...,10',
                                    '0:10:1')
        if input_str is None:
            return
        station_numbers = []
        parts = input_str.split(';')
        for part in parts:
            part = part.strip()
            if not part:
                continue
            if ':' in part:
                try:
                    segs = [float(x) for x in part.split(':')]
                    if len(segs) == 3:
                        station_numbers.extend(core.make_seq_prealloc(segs[0], segs[1], segs[2]))
                    elif len(segs) == 2:
                        station_numbers.extend(np.arange(segs[0], segs[1] + 1, 1.0))
                except ValueError:
                    messagebox.showerror('输入错误', '区间格式错误: %s' % part, parent=self.root)
                    return
            else:
                for tok in part.split(','):
                    tok = tok.strip()
                    if tok:
                        try:
                            station_numbers.append(float(tok))
                        except ValueError:
                            messagebox.showerror('输入错误', '站号格式错误: %s' % tok, parent=self.root)
                            return
        if not station_numbers:
            messagebox.showerror('输入错误', '未生成任何有效站号。', parent=self.root)
            return
        self._save_undo('Half_table', '快速站号')
        rows = []
        for i, s in enumerate(station_numbers):
            rows.append([i + 1, core.num2trimstr(s), '', '', ''])
        self.Half_table.set_columns(['列', '站号', '半宽', '系数', '相对矩臂'])
        self.Half_table.set_data(rows)
        self.log('已生成 %d 行站号数据。' % len(rows))
        self.update_half_width_plot()

    def add_half_clicked(self):
        """快速半宽导入（粘贴 Excel 单列 + 单位换算）"""
        station_col = self.Half_table.get_data_as_columns().get('站号', [])
        station_num = np.array([core._parse_scalar(v) for v in station_col], dtype=float)
        mask = np.isfinite(station_num)
        n_need = int(np.sum(mask))
        if n_need <= 0:
            messagebox.showerror('提示', '站号列没有有效数据，请先生成站号。', parent=self.root)
            return
        text = ask_multiline_input(self.root, '导入半宽（支持从Excel粘贴）',
                                   '请粘贴 %d 行半宽数据（与有效站号行数一致）：' % n_need)
        if text is None:
            return
        lines = [l for l in text.splitlines() if l.strip() != '']
        try:
            vals = core.parse_pasted_numbers(lines)
        except ValueError as e:
            messagebox.showerror('错误', str(e), parent=self.root)
            return
        if len(vals) != n_need:
            messagebox.showerror('数量不匹配',
                                 '粘贴数量 (%d) 与有效站号行数 (%d) 不一致。' % (len(vals), n_need),
                                 parent=self.root)
            return
        from src.ui.ui_widgets import ask_text_dialog
        unit = ask_text_dialog(self.root, '单位选择', '粘贴数据单位 (m/cm/mm/custom):', 'm')
        scale = 1.0
        if unit is not None:
            u = unit.lower().strip()
            if u == 'cm':
                scale = 0.01
            elif u == 'mm':
                scale = 0.001
            elif u == 'custom':
                ans = ask_text_dialog(self.root, '自定义比例', '每单位等于多少米:', '1')
                try:
                    scale = float(ans) if ans else 1.0
                except ValueError:
                    scale = 1.0
        vals_m = vals * scale
        self._save_undo('Half_table', '快速半宽导入')
        idx = np.nonzero(mask)[0]
        data = self.Half_table.get_data()
        for k, r in enumerate(idx):
            r = int(r)
            while len(data[r]) < 5:
                data[r].append('')
            data[r][3] = core.num2trimstr(vals_m[k])  # 半宽列（第3列，0基）
        self.Half_table.set_data(data)
        self.log('半宽导入完成：%d 行（单位已转换为 m）。' % n_need)
        self.update_half_width_plot()

    def add_coefficient_clicked(self):
        """添加积分系数列（梯形/Simpson 1/3 / 3/8）"""
        cols = self.Half_table.get_data_as_columns()
        station_raw = cols.get('站号', [])
        station = np.array([core._parse_scalar(v) for v in station_raw], dtype=float)
        if np.any(np.isnan(station)):
            messagebox.showerror('错误', '"站号"列存在空白/非法值，请先修正。', parent=self.root)
            return
        if np.any(np.diff(station) < 0):
            messagebox.showerror('错误', '"站号"应非降序排列，请先排序。', parent=self.root)
            return
        n = station.size
        if n < 2:
            messagebox.showerror('警告', '至少需要两个站号。', parent=self.root)
            return
        # 分段检测
        segs = core.detect_segments_by_station(station, 1e-9)
        if not segs:
            segs = [[1, n, math.nan]]
        self.StationSegments = np.array(segs)
        coeff = np.zeros(n)
        method = self.CoefficientMethod
        for seg in segs:
            a = int(seg[0]) - 1  # 0基
            b = int(seg[1]) - 1
            m = b - a + 1
            if m <= 1:
                continue
            if method == 'trapezoidal':
                coeff[a] += 0.5
                if m > 2:
                    coeff[a + 1:b] += 1
                coeff[b] += 0.5
            elif method == 'simp1':
                r = (m - 1) % 2
                last = b - r
                if last - a >= 2:
                    coeff[a] += 1
                    idx = list(range(a + 1, last))
                    for j, ii in enumerate(idx):
                        coeff[ii] += 4 if (ii - a) % 2 == 1 else 2
                    coeff[last] += 1
                if r == 1:
                    coeff[last] += 0.5
                    coeff[b] += 0.5
            elif method == 'simp2':
                q = (m - 1) // 3
                covered = a + 3 * q
                p = a
                for _ in range(q):
                    coeff[p] += 1
                    coeff[p + 1] += 3
                    coeff[p + 2] += 3
                    coeff[p + 3] += 1
                    p += 3
                rem = (m - 1) - 3 * q
                if rem == 2:
                    coeff[covered] += 1
                    coeff[covered + 1] += 4
                    coeff[b] += 1
                elif rem == 1:
                    coeff[covered] += 0.5
                    coeff[b] += 0.5
            else:
                messagebox.showerror('警告', '请先选择系数方法。', parent=self.root)
                return
        self._save_undo('Half_table', '添加系数')
        data = self.Half_table.get_data()
        for i in range(n):
            if len(data[i]) < 4:
                data[i] = list(data[i]) + [''] * (4 - len(data[i]))
            data[i][3] = core.num2trimstr(coeff[i])
        self.Half_table.set_data(data)
        nseg = sum(1 for s in segs if s[1] > s[0])
        self.var_Segment.set(str(nseg))
        msg = {'trapezoidal': '已分段（等距段）按梯形法生成系数。',
               'simp1': '已分段（等距段）按辛普森1/3混合法生成系数。',
               'simp2': '已分段（等距段）按辛普森3/8混合法生成系数。'}
        self.log(msg.get(method, '已生成系数。'))
        messagebox.showinfo('成功', msg.get(method, '已生成系数。'), parent=self.root)

    def add_moment_arm_clicked(self):
        """添加相对矩臂列"""
        cols = self.Half_table.get_data_as_columns()
        station_raw = cols.get('站号', [])
        station = np.array([core._parse_scalar(v) for v in station_raw], dtype=float)
        finite = station[np.isfinite(station)]
        if finite.size == 0:
            messagebox.showerror('提示', '站号列为空，请先生成站号。', parent=self.root)
            return
        if self.OriginFlag == 'amidship':
            origin = (np.min(finite) + np.max(finite)) / 2
        elif self.OriginFlag == 'stern':
            origin = np.min(finite)
        elif self.OriginFlag == 'bow':
            origin = np.max(finite)
        else:
            origin = (np.min(finite) + np.max(finite)) / 2
        ans = ask_text_dialog(self.root, '原点设置', '原点站号（自动计算为 %.4f）:' % origin,
                              core.num2trimstr(origin))
        if ans is None:
            return
        try:
            origin = float(ans)
        except ValueError:
            pass
        arms = station - origin
        self._save_undo('Half_table', '添加相对矩臂')
        data = self.Half_table.get_data()
        for i in range(station.size):
            if len(data[i]) < 5:
                data[i] = list(data[i]) + [''] * (5 - len(data[i]))
            data[i][4] = core.num2trimstr(arms[i])
        self.Half_table.set_data(data)
        self.log('相对矩臂列已生成。')

    def add_row_clicked(self):
        self._save_undo('Half_table', '新增行')
        n = self.Half_table.column_count()
        self.Half_table.add_row([self.Half_table.row_count() + 1] + [''] * (n - 1))
        self.update_half_width_plot()

    def delete_row_clicked(self):
        self._save_undo('Half_table', '删除行')
        self.Half_table.delete_last_row()
        data = self.Half_table.get_data()
        for i, row in enumerate(data):
            if row:
                row[0] = i + 1
        self.Half_table.set_data(data)
        self.update_half_width_plot()

    def lock_edit_clicked(self):
        self.IsLocked = not self.IsLocked
        self.Half_table.set_editable(not self.IsLocked)
        self._update_button_state()
        self._refresh_statusbar()
        self.log('表格已%s编辑。' % ('锁定' if self.IsLocked else '解锁'))

    def delete_col_clicked(self):
        if not messagebox.askyesno('确认删除', '数值显示框归零，是否继续？', parent=self.root):
            return
        for v in ['var_A', 'var_M', 'var_Half_A', 'var_Half_M', 'var_Ful_A',
                  'var_Full_M', 'var_LCF', 'var_Segment']:
            getattr(self, v).set('')
        self.log('数值显示框已归零。')

    def symmetry_clicked(self):
        self.IsSymmetricView = not self.IsSymmetricView
        self.update_half_width_plot()
        self.update_transverse_section_plot()

    def subsection_clicked(self):
        """分段数据显示"""
        if not self._check_principal():
            return
        cols = self.Half_table.get_data_as_columns()
        station = np.array([core._parse_scalar(v) for v in cols.get('站号', [])], dtype=float)
        half = np.array([core._parse_scalar(v) for v in cols.get('半宽', [])], dtype=float)
        coeff = np.array([core._parse_scalar(v) for v in cols.get('系数', [])], dtype=float)
        arm = np.array([core._parse_scalar(v) for v in cols.get('相对矩臂', [])], dtype=float)
        if np.any(~np.isfinite(station)) or np.any(~np.isfinite(half)):
            messagebox.showerror('数据不完整', '数据中存在空白或非法值（NaN/Inf）。', parent=self.root)
            return
        segs = self.StationSegments
        if segs is None:
            segs = core.detect_segments_by_station(station, 1e-9)
        if segs is None or len(segs) == 0:
            segs = [[1, len(station), math.nan]]
        segs = np.array(segs)
        n = len(station)
        segs_clean = core.preprocess_segments(segs, n)
        if segs_clean.shape[0] == 0:
            segs_clean = np.array([[1, n, np.nan, np.nan]])
        scale_global, _ = core.method_scale(self.CoefficientMethod)
        total_span = self.LppEndStation - self.LppStartStation
        lines = []
        lines.append('分段统计:')
        for k in range(segs_clean.shape[0]):
            i0 = int(segs_clean[k, 0]) - 1
            i1 = int(segs_clean[k, 1]) - 1
            if i1 <= i0:
                continue
            station_range = station[i1] - station[i0]
            point_count = i1 - i0
            h_station = station_range / point_count if point_count > 0 else 0
            h_len = h_station * (self.Lpp / total_span) if total_span > 0 else 0
            scale_k = segs_clean[k, 3] if math.isfinite(segs_clean[k, 3]) and segs_clean[k, 3] > 0 else scale_global
            idx = np.arange(i0, i1 + 1)
            sum_ac = float(np.sum(half[idx] * coeff[idx]))
            sum_am = float(np.sum(half[idx] * coeff[idx] * arm[idx]))
            half_a = scale_k * h_len * sum_ac
            half_m = scale_k * h_len * h_len * sum_am
            lines.append('段#%d: 行%d-%d, 点数=%d, 步长=%.4f, Half_A=%.4f, Full_A=%.4f, Half_M=%.4f, Full_M=%.4f'
                         % (k + 1, i0 + 1, i1 + 1, point_count + 1, h_len,
                            half_a, 2 * half_a, half_m, 2 * half_m))
        nseg = segs_clean.shape[0]
        self.var_Segment.set(str(nseg))
        self._info_dialog('分段统计', lines)

    def cal_clicked(self):
        """水线面核心计算（分段累加）"""
        if not self._check_principal():
            return
        cols = self.Half_table.get_data_as_columns()
        station = np.array([core._parse_scalar(v) for v in cols.get('站号', [])], dtype=float)
        half = np.array([core._parse_scalar(v) for v in cols.get('半宽', [])], dtype=float)
        coeff = np.array([core._parse_scalar(v) for v in cols.get('系数', [])], dtype=float)
        arm = np.array([core._parse_scalar(v) for v in cols.get('相对矩臂', [])], dtype=float)
        if self.Half_table.row_count() < 2:
            messagebox.showerror('提示', '表格至少需要两行数据！', parent=self.root)
            return
        if any(c not in cols for c in ['站号', '半宽', '系数', '相对矩臂']):
            messagebox.showerror('错误', '缺少"站号"、"半宽"、"系数"或"相对矩臂"列！', parent=self.root)
            return
        if np.any(~np.isfinite(station)) or np.any(~np.isfinite(half)) or \
           np.any(~np.isfinite(coeff)) or np.any(~np.isfinite(arm)):
            messagebox.showerror('数据不完整', '数据中存在空白或非法值，请先修正。', parent=self.root)
            return
        n = len(station)
        scale_global, method_name = core.method_scale(self.CoefficientMethod)
        self.log('使用积分方法：%s' % method_name)
        # 分段
        if self.StationSegments is not None:
            segs = core.preprocess_segments(self.StationSegments, n)
            if segs.shape[0] == 0:
                segs = None
        else:
            segs = None
        if segs is None:
            auto = core.detect_segments_by_station(station, 1e-9)
            if auto:
                segs = np.array([[s[0], s[1], s[2], np.nan] for s in auto], dtype=float)
            else:
                segs = np.array([[1, n, np.nan, np.nan]], dtype=float)
        res = core.calc_waterplane_segments(station, half, coeff, arm, segs,
                                            scale_global, self.Lpp,
                                            self.LppStartStation, self.LppEndStation)
        for d in res['debug']:
            self.log(d)
        self.var_A.set(core.num2trimstr(float(np.sum(half * coeff)), 1e-9))
        self.var_M.set(core.num2trimstr(float(np.sum(half * coeff * arm)), 1e-9))
        self.var_Half_A.set('%.4f' % res['half_a'])
        self.var_Half_M.set('%.4f' % res['half_m'])
        self.var_Ful_A.set('%.4f' % res['full_a'])
        self.var_Full_M.set('%.4f' % res['full_m'])
        if math.isfinite(res['lcf']):
            self.var_LCF.set('%.4f' % res['lcf'])
        else:
            self.var_LCF.set('')
            messagebox.showwarning('警告', '水线面面积为零，无法计算漂心纵坐标。', parent=self.root)
        self.log('计算完成: half_A=%.4f, half_M=%.4f, full_A=%.4f, full_M=%.4f, LCF=%.4f'
                 % (res['half_a'], res['half_m'], res['full_a'], res['full_m'],
                    res['lcf'] if math.isfinite(res['lcf']) else 0))
        messagebox.showinfo('计算成功', '数值计算完成！', parent=self.root)

    def _check_principal(self):
        if not math.isfinite(self.Lpp) or self.Lpp <= 0:
            messagebox.showerror('主尺度缺失', '请先设置垂线间长 Lpp ！', parent=self.root)
            return False
        if not math.isfinite(self.LppStartStation) or not math.isfinite(self.LppEndStation):
            messagebox.showerror('主尺度缺失', '请先设置垂线间长的始末站号！', parent=self.root)
            return False
        if self.LppEndStation <= self.LppStartStation:
            messagebox.showerror('主尺度设置错误', '垂线间长结束站号必须大于起始站号！', parent=self.root)
            return False
        return True

    # =====================================================================
    # 曲线拟合
    # =====================================================================
    def curve_fitting_clicked(self):
        if not self._check_principal():
            return
        scale = self.Lpp / (self.LppEndStation - self.LppStartStation)
        cols = self.Half_table.get_data_as_columns()
        station_raw = cols.get('站号', [])
        half_raw = cols.get('半宽', [])
        station = np.array([core._parse_scalar(v) for v in station_raw], dtype=float)
        half = np.array([core._parse_scalar(v) for v in half_raw], dtype=float)
        good = np.isfinite(station) & np.isfinite(half)
        station = station[good]
        half = half[good]
        if station.size < 3:
            messagebox.showerror('数据不足', '可用数据点不足（<3），无法拟合。', parent=self.root)
            return
        x_raw = (station - self.LppStartStation) * scale
        from src.ui.ui_widgets import ask_text_dialog
        cfg = ask_text_dialog(
            self.root, '曲线拟合配置',
            '模型 (poly/pchip):\n多项式最大阶数(1-6):\n自动选阶(yes/no):\n固定阶数:\n'
            '纵向范围x(米, 如 -5:50 留空=数据范围):\n采样点数(>=100):\n去除离群(yes/no):',
            'poly;4;yes;3;;400;yes')
        if cfg is None:
            return
        parts = [p.strip() for p in cfg.split(';')]
        model = parts[0].lower() if parts and parts[0] else 'poly'
        if model not in ('poly', 'pchip'):
            model = 'poly'
        max_deg = int(float(parts[1])) if len(parts) > 1 and parts[1] else 4
        auto_deg = parts[2].lower() in ('yes', 'y', 'true', '是') if len(parts) > 2 else True
        fix_deg = int(float(parts[3])) if len(parts) > 3 and parts[3] else 3
        range_str = parts[4] if len(parts) > 4 else ''
        n_pts = int(float(parts[5])) if len(parts) > 5 and parts[5] else 400
        rm_out = parts[6].lower() in ('yes', 'y', 'true', '是') if len(parts) > 6 else True
        use_range = False
        x_min = float(np.min(x_raw))
        x_max = float(np.max(x_raw))
        if range_str and ':' in range_str:
            try:
                r1, r2 = [float(t) for t in range_str.split(':')[:2]]
                if r2 >= r1:
                    use_range = True
                    x_min, x_max = r1, r2
            except ValueError:
                pass
        x = x_raw
        y = half
        if use_range:
            m = (x >= x_min) & (x <= x_max)
            x, y = x[m], y[m]
        if rm_out:
            try:
                from scipy import stats
                keep = ~(np.abs(stats.zscore(x)) > 3) & ~(np.abs(stats.zscore(y)) > 3)
                x, y = x[keep], y[keep]
            except Exception:
                pass
        if x.size < 3:
            messagebox.showerror('数据不足', '剔除后有效数据点不足，无法拟合。', parent=self.root)
            return
        # 去重 + 排序
        uniq_x, idx = np.unique(x, return_index=True)
        x, y = uniq_x, y[idx]
        order = np.argsort(x)
        x, y = x[order], y[order]
        ax = self.plot_fitting
        lines = ['曲线拟合配置']
        try:
            if model == 'poly':
                max_allowed = max(1, min([max_deg, x.size - 1, 6]))
                best_deg, best_rmse, best_coeff = 1, math.inf, None
                if auto_deg:
                    for d in range(1, max_allowed + 1):
                        if x.size <= d:
                            break
                        c, rmse = core.polyfit_rmse(x, y, d)
                        if rmse < best_rmse:
                            best_rmse, best_deg, best_coeff = rmse, d, c
                    deg = best_deg
                    coeff = best_coeff
                else:
                    deg = min(fix_deg, max_allowed)
                    coeff, best_rmse = core.polyfit_rmse(x, y, deg)
                yhat = np.polyval(coeff, x)
                x_fit = np.linspace(x_min, x_max, n_pts)
                y_fit = np.polyval(coeff, x_fit)
                f_expr = core.poly_expr(coeff)
                model_name = 'Poly(%d)' % deg
                lines.append('模型: POLY, 自动阶数: %s, 阶数: %d' % (auto_deg, deg))
            else:
                from scipy.interpolate import PchipInterpolator
                pp = PchipInterpolator(x, y)
                yhat = pp(x)
                best_rmse = float(np.sqrt(np.mean((y - yhat) ** 2)))
                x_fit = np.linspace(x_min, x_max, n_pts)
                y_fit = pp(x_fit)
                f_expr = 'f(x) = pchip interpolation on %d breakpoints' % x.size
                model_name = 'PCHIP'
                lines.append('模型: PCHIP')
            res = y - yhat
            mae = float(np.mean(np.abs(res)))
            max_abs = float(np.max(np.abs(res)))
            sst = float(np.sum((y - np.mean(y)) ** 2))
            sse = float(np.sum(res ** 2))
            r2 = 1 - sse / sst if sst > 0 else math.nan
            r_xy = float(np.corrcoef(x, y)[0, 1]) if x.size > 1 else math.nan
            r_yy = float(np.corrcoef(y, yhat)[0, 1]) if y.size > 1 else math.nan
            # 绘图
            ax.clear()
            ax.ax.plot(x_raw, half, 'o', color='#a6a6a6', markersize=5, label='原始数据')
            ax.ax.plot(x, y, 'bo', markersize=6, label='有效数据')
            ax.ax.plot(x_fit, y_fit, 'r-', linewidth=2.2,
                       label='拟合曲线 [%s], RMSE=%.4g' % (model_name, best_rmse))
            ax.ax.axhline(0, color='k', linewidth=0.8)
            ax.ax.axvline(0, color='k', linewidth=0.8)
            ax.ax.set_xlabel('纵向位置 x (m)')
            ax.ax.set_ylabel('半宽 (m)')
            ax.ax.set_title('曲线拟合: %s (有效点=%d)' % (model_name, x.size))
            ax.ax.legend(loc='best', fontsize=8)
            ax.ax.grid(True, alpha=0.3)
            ax.refresh()
            lines.append('拟合结果: %s, 有效样本数: %d' % (model_name, x.size))
            lines.append('RMSE: %.6e, MAE: %.6e, Max|误差|: %.6e' % (best_rmse, mae, max_abs))
            lines.append('R²: %.6f' % r2)
            lines.append('相关系数 r(x,y): %.6f, r(y,ŷ): %.6f' % (r_xy, r_yy))
            lines.append('')
            lines.append('函数表达式 f(x)  (x 单位: 米)')
            lines.append('  ' + f_expr)
        except Exception as e:
            messagebox.showerror('曲线拟合失败', '拟合失败：%s' % e, parent=self.root)
            return
        self.TextArea_curve_fitting.delete('1.0', 'end')
        self.TextArea_curve_fitting.insert('1.0', '\n'.join(lines))

    # =====================================================================
    # 水线面绘图
    # =====================================================================
    def update_half_width_plot(self):
        """绘制半宽水线面曲线图"""
        ax = self.plot_half_area
        if not self._principal_ok():
            ax.show_message('请先设置垂线间长和始末站号', '纵向位置 (m)', '横向位置 (m)')
            return
        cols = self.Half_table.get_data_as_columns()
        station_col = cols.get('站号', [])
        half_col = cols.get('半宽', [])
        station = np.array([core._parse_scalar(v) for v in station_col], dtype=float)
        half = np.array([core._parse_scalar(v) for v in half_col], dtype=float)
        valid = np.isfinite(station) & np.isfinite(half)
        if np.sum(valid) < 1:
            ax.show_message('无有效数据', '纵向位置 (m)', '横向位置 (m)')
            return
        stations = station[valid]
        half_widths = half[valid]
        ratio = self.Lpp / (self.LppEndStation - self.LppStartStation)
        x = (stations - self.LppStartStation) * ratio
        order = np.argsort(x)
        x, half_widths = x[order], half_widths[order]
        within = (x >= 0) & (x <= self.Lpp)
        ax.clear()
        if self.IsSymmetricView:
            if x.size > 2:
                xf = np.concatenate([x, x[::-1]])
                yf = np.concatenate([half_widths, -half_widths[::-1]])
                ax.ax.fill(xf, yf, color='#b3d9ff', alpha=0.6)
            ax.ax.plot(x, half_widths, 'b-', linewidth=2.5, label='右舷轮廓')
            ax.ax.plot(x, -half_widths, 'r-', linewidth=2.5, label='左舷轮廓')
            ax.ax.plot(x[within], half_widths[within], 'bo', markersize=4)
            ax.ax.plot(x[within], -half_widths[within], 'ro', markersize=4)
            ymax = float(np.max(half_widths))
            ylim = [-ymax * 1.1, ymax * 1.1]
        else:
            ax.ax.plot(x, half_widths, 'b-', linewidth=2.5, label='半宽曲线')
            ax.ax.plot(x[within], half_widths[within], 'ro', markersize=5)
            outside = ~within
            if np.any(outside):
                ax.ax.plot(x[outside], half_widths[outside], 'o', markersize=5,
                           color='#ff9900')
            ylim = [0, float(np.max(half_widths)) * 1.1]
        xlim = [float(np.min(x)), float(np.max(x))]
        if x.size > 1:
            ax.ax.plot(xlim, [0, 0], 'k-', linewidth=1.5, label='船体中线')
        ax.ax.plot([0, 0], ylim, 'g-', linewidth=1.5, label='垂线')
        ax.ax.plot([self.Lpp, self.Lpp], ylim, 'g-', linewidth=1.5)
        ax.ax.set_xlabel('纵向位置 (m)')
        ax.ax.set_ylabel('横向位置 (m)')
        ax.ax.set_xlim(xlim)
        ax.ax.set_ylim(ylim)
        # 等比例下图形很扁，图例横排放到图外下方，避免遮挡型线
        ax.ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.14),
                     ncol=3, fontsize=7, frameon=False)
        ax.ax.grid(True, alpha=0.3)
        # 船长（~200 m）与半宽（~6 m）按真实长度比例显示
        ax.set_true_aspect()
        ax.refresh()
        # 更新水线面面积
        if x.size >= 3:
            half_area = float(np.trapezoid(half_widths, x))
            full_area = 2 * half_area
            moment = float(np.trapezoid(x * half_widths, x))
            lcf = moment / half_area if abs(half_area) > 1e-12 else math.nan
            self.var_Half_A.set('%.4f' % half_area)
            self.var_Ful_A.set('%.4f' % full_area)
            if math.isfinite(lcf):
                self.var_LCF.set('%.4f' % lcf)

    def _principal_ok(self):
        return (math.isfinite(self.Lpp) and self.Lpp > 0 and
                math.isfinite(self.LppStartStation) and math.isfinite(self.LppEndStation)
                and self.LppEndStation > self.LppStartStation)

    # =====================================================================
    # 横剖面
    # =====================================================================
    def update_transverse_section_plot(self):
        ax = self.plot_z_area
        cols = self.Z_table.get_data_as_columns()
        height = np.array([core._parse_scalar(v) for v in cols.get('高度', [])], dtype=float)
        half = np.array([core._parse_scalar(v) for v in cols.get('半宽', [])], dtype=float)
        valid = np.isfinite(height) & np.isfinite(half)
        if np.sum(valid) < 2:
            ax.show_message('横剖面图 (有效数据不足)', '横向位置 (m)', '垂向高度 (m)')
            return
        heights = height[valid]
        half_widths = half[valid]
        order = np.argsort(heights)
        heights, half_widths = heights[order], half_widths[order]
        ax.clear()
        if self.IsSymmetricView:
            xf = np.concatenate([half_widths, -half_widths[::-1]])
            yf = np.concatenate([heights, heights[::-1]])
            ax.ax.fill(xf, yf, color='#b3d9ff', alpha=0.6)
            ax.ax.plot(half_widths, heights, 'b-', linewidth=2, label='右舷')
            ax.ax.plot(-half_widths, heights, 'r-', linewidth=2, label='左舷')
            xlim = [-float(np.max(half_widths)) * 1.2, float(np.max(half_widths)) * 1.2]
        else:
            ax.ax.plot(half_widths, heights, 'b-', linewidth=2.5, label='半宽曲线')
            ax.ax.plot(half_widths, heights, 'ro', markersize=5, label='数据点')
            xlim = [0, float(np.max(half_widths)) * 1.2]
        ymin, ymax = float(np.min(heights)), float(np.max(heights))
        yrange = (ymax - ymin) or 1.0
        ylim = [ymin - 0.1 * yrange, ymax + 0.1 * yrange]
        if ylim[0] > 0:
            ylim[0] = 0
        ax.ax.plot([0, 0], ylim, 'k-', linewidth=2, label='中心线')
        ax.ax.plot(xlim, [0, 0], 'k-', linewidth=1, label='基线')
        ax.ax.set_xlabel('横向位置 (m)')
        ax.ax.set_ylabel('垂向高度 (m)')
        ax.ax.set_xlim(xlim)
        ax.ax.set_ylim(ylim)
        ax.ax.grid(True, alpha=0.3)
        ax.ax.legend(loc='best', fontsize=8)
        # 半宽与高度同为长度量纲，按真实比例显示
        ax.set_true_aspect()
        ax.refresh()

    def calc_transverse_section_clicked(self):
        if not self._check_principal():
            return
        cols = self.Z_table.get_data_as_columns()
        height = np.array([core._parse_scalar(v) for v in cols.get('高度', [])], dtype=float)
        half = np.array([core._parse_scalar(v) for v in cols.get('半宽', [])], dtype=float)
        coeff_raw = cols.get('系数', [])
        coeff = np.array([core._parse_scalar(v) for v in coeff_raw], dtype=float)
        if len(height) == 0 or len(half) == 0:
            messagebox.showerror('错误', '缺少"高度"或"半宽"列！', parent=self.root)
            return
        if '系数' not in cols or np.all(np.isnan(coeff)):
            coeff = np.ones(len(height))
        valid = np.isfinite(height) & np.isfinite(half) & np.isfinite(coeff)
        if np.sum(valid) < 2:
            messagebox.showerror('数据不完整', '有效数据不足，至少需要两个有效数据点。', parent=self.root)
            return
        res = core.calc_transverse_section(height[valid], half[valid], coeff[valid],
                                           self.CoefficientMethod)
        self.var_Z_HalfArea.set('%.4f' % res['halfArea'])
        self.var_Z_FullArea.set('%.4f' % res['fullArea'])
        self.var_Z_CentroidZ.set('%.4f' % res['centroid_z'])
        self.log('半船面积 = %.4f m²' % res['halfArea'])
        self.log('全船面积 = %.4f m²' % res['fullArea'])
        self.log('垂向形心 z_c = %.4f m' % res['centroid_z'])
        # 更新横剖面图标记形心
        self.update_transverse_section_plot()
        # 存储结果到 SectionAreas
        station_num = self._current_station_num()
        if station_num is not None:
            spacing = self.Lpp / (self.LppEndStation - self.LppStartStation)
            mid = (self.LppStartStation + self.LppEndStation) / 2
            x_pos = (station_num - mid) * spacing
            self.var_Z_HalfCentroidX.set('%.3f' % x_pos)
            sl = self.SectionAreas
            if station_num not in sl['stations_list']:
                sl['stations_list'].append(station_num)
                sl['halfAreas_list'].append(res['halfArea'])
                sl['fullAreas_list'].append(res['fullArea'])
                sl['centroids_y_list'].append(res['half_centroid_y'])
                sl['centroids_z_list'].append(res['centroid_z'])
                sl['station_positions_list'].append(x_pos)
            else:
                i = sl['stations_list'].index(station_num)
                sl['halfAreas_list'][i] = res['halfArea']
                sl['fullAreas_list'][i] = res['fullArea']
                sl['centroids_y_list'][i] = res['half_centroid_y']
                sl['centroids_z_list'][i] = res['centroid_z']
                sl['station_positions_list'][i] = x_pos
            # 排序
            order = np.argsort(sl['stations_list'])
            for k in sl:
                sl[k] = list(np.asarray(sl[k])[order])
        messagebox.showinfo('计算完成', '横剖面形心计算完成！', parent=self.root)

    def _current_station_num(self):
        sel = self._selected_node()
        if not sel:
            return None
        meta = self.tree_meta.get(sel, {})
        if meta.get('type') == 'bodyplan':
            return meta.get('station')
        text = self.tree.item(sel, 'text')
        import re
        m = re.search(r'站\s*(\d+\.?\d*)', str(text))
        if m:
            return float(m.group(1))
        return None

    # =====================================================================
    # 导出 / 项目保存
    # =====================================================================
    def menu_export(self):
        """导出表格数据（CSV / Excel / TXT），原子写入避免半截文件。

        按当前 Tab 导出对应表格；xlsx 一次导出 半宽表/横剖面表/原表格 三个工作表。
        """
        if self.isTap == 1 and self.original_headers:
            cols, rows, title = self.original_headers, self.original_data, '原表格'
        elif self.isTap == 3:
            cols, rows, title = self.Z_table.get_columns(), self.Z_table.get_data(), '横剖面'
        else:
            cols, rows, title = self.Half_table.get_columns(), self.Half_table.get_data(), '半宽'
        path = filedialog.asksaveasfilename(
            title='保存表格数据（%s）' % title,
            filetypes=[('CSV 文件', '*.csv'), ('Excel 文件', '*.xlsx'),
                       ('文本文件', '*.txt')],
            defaultextension='.csv', parent=self.root)
        if not path:
            return
        import tempfile
        d = os.path.dirname(os.path.abspath(path))
        tmp = os.path.join(d, '.' + os.path.basename(path) + '.tmp')
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext == '.xlsx':
                import openpyxl
                wb = openpyxl.Workbook()
                sheets = [('半宽表', self.Half_table.get_columns(), self.Half_table.get_data()),
                          ('横剖面表', self.Z_table.get_columns(), self.Z_table.get_data())]
                if self.original_headers:
                    sheets.append(('原表格', self.original_headers, self.original_data))
                first = True
                for sname, hcols, hrows in sheets:
                    ws = wb.active if first else wb.create_sheet()
                    first = False
                    ws.title = sname[:31]
                    ws.append([str(c) for c in hcols])
                    for r in hrows:
                        ws.append([EditableTable._fmt(v) for v in r])
                wb.save(tmp)
            else:
                import csv
                sep = '\t' if ext == '.txt' else ','
                with open(tmp, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f, delimiter=sep)
                    writer.writerow([str(c) for c in cols])
                    for row in rows:
                        writer.writerow([EditableTable._fmt(v) for v in row])
            os.replace(tmp, path)
            self.log('表格数据已导出到: %s' % path)
            messagebox.showinfo('导出成功', '数据已成功导出到: %s' % path, parent=self.root)
        except Exception as e:
            try:
                if os.path.exists(tmp):
                    os.remove(tmp)
            except OSError:
                pass
            messagebox.showerror('导出失败', '导出失败: %s' % e, parent=self.root)

    def _embed_project_logo(self):
        """把应用 logo（icon/船.png）以 base64 内嵌进项目文件。"""
        import base64
        try:
            icon_dir = getattr(self, 'icon_dir', None) or 'icon'
            for name in ('船.png', 'ship.png'):
                p = os.path.join(icon_dir, name)
                if os.path.exists(p):
                    with open(p, 'rb') as f:
                        return base64.b64encode(f.read()).decode('ascii')
        except Exception:
            pass
        return None

    @staticmethod
    def _atomic_save(path, obj):
        """原子写 pickle 文件：写临时文件 → fsync → 回读校验 → os.replace。

        避免程序中途崩溃/磁盘写满导致原文件损坏；校验通过才覆盖正式文件。
        """
        import tempfile
        import pickle as _pickle
        d = os.path.dirname(os.path.abspath(path))
        base = os.path.basename(path)
        fd, tmp = tempfile.mkstemp(suffix='.tmp', prefix=base + '.', dir=d)
        try:
            with os.fdopen(fd, 'wb') as f:
                _pickle.dump(obj, f, protocol=_pickle.HIGHEST_PROTOCOL)
                f.flush()
                os.fsync(f.fileno())
            with open(tmp, 'rb') as f:      # 回读校验：确保文件可反序列化
                back = _pickle.load(f)
            if back is None:
                raise RuntimeError('校验失败：回读得到空对象')
            os.replace(tmp, path)
            return back
        finally:
            if os.path.exists(tmp):
                try:
                    os.remove(tmp)
                except OSError:
                    pass

    def _build_project_payload(self):
        """收集当前项目全部状态（版本化），供保存使用"""
        ml_model = self.ML_model
        if ml_model is not None:
            try:
                import pickle as _pickle
                _pickle.dumps(ml_model)     # 可序列化才保存，否则置空不阻塞保存
            except Exception:
                ml_model = None
        return dict(
            version=3,
            logo=self._embed_project_logo(),
            principal=dict(Lpp=self.Lpp, Breadth=self.Breadth, Depth=self.Depth,
                           LppStartStation=self.LppStartStation,
                           LppEndStation=self.LppEndStation,
                           Draft=self.Draft, HeelAngle=self.HeelAngle,
                           TrimAngle=self.TrimAngle, OriginFlag=self.OriginFlag,
                           CoefficientMethod=self.CoefficientMethod),
            waterlines=self.waterlines, decklines=self.decklines,
            bodyplans=self.bodyplans, sections=self.sections,
            half_table_data=self.Half_table.get_data(),
            half_table_cols=self.Half_table.get_columns(),
            z_table_data=self.Z_table.get_data(),
            z_table_cols=self.Z_table.get_columns(),
            original_data=self.original_data, original_headers=self.original_headers,
            hydrostatics=self.Hydrostatics, bonjean=self.BonjeanCurves,
            stability=self.StabilityData, gz=self.GZ_CurveData,
            dynamic=self.DynamicStabilityData,
            buoyancy=dict(volume=self.BuoyancyVolume, center=self.BuoyancyCenter),
            section_areas=self.SectionAreas,
            ml_model=ml_model,
            ui_state=dict(IsLocked=self.IsLocked, IsSymmetricView=self.IsSymmetricView,
                          WireframeMode=self.WireframeMode, SurfaceColor=self.SurfaceColor))

    def _open_project_file(self, path, from_recent=False, add_to_recent=True):
        """打开 .scs 项目文件并恢复数据；成功返回 True。"""
        if not os.path.exists(path):
            if from_recent:
                self._remove_recent_project(path)
                messagebox.showwarning('文件不存在',
                                       '最近文件已不存在或已被移动:\n%s' % path, parent=self.root)
            else:
                messagebox.showerror('打开失败', '文件不存在: %s' % path, parent=self.root)
            return False
        try:
            import pickle
            with open(path, 'rb') as f:
                p = pickle.load(f)
            if not isinstance(p, dict) or 'principal' not in p:
                raise RuntimeError('不是有效的 SCS 项目文件。')
            self._apply_project_payload(p)
            self._apply_project_logo(p)
            self._clear_dirty()
            self._current_project_path = os.path.abspath(path)
            if add_to_recent:
                self._add_recent_project(path)
            self._refresh_statusbar()
            if from_recent:
                self.log('项目已打开: %s' % path)
            else:
                self.log('项目已成功导入并重建。')
                messagebox.showinfo('导入成功', '项目已成功导入！', parent=self.root)
            return True
        except Exception as e:
            messagebox.showerror('导入失败', '导入失败: %s' % e, parent=self.root)
            return False

    def _open_recent(self, path):
        """从"最近打开"菜单打开项目。"""
        self._open_project_file(path, from_recent=True)

    # ---------- 自动保存 ----------

    def _autosave_backup_path(self):
        """自动保存备份路径：已命名项目→同目录 .autosave.scs；未命名→用户目录。"""
        try:
            cur = getattr(self, '_current_project_path', None)
            if cur:
                d = os.path.dirname(cur)
                b = os.path.basename(cur)
                base = b[:-4] if b.lower().endswith('.scs') else b
                return os.path.join(d, base + '.autosave.scs')
            return os.path.join(os.path.expanduser('~'), 'scs_autosave.scs')
        except Exception:
            return os.path.join(os.path.expanduser('~'), 'scs_autosave.scs')

    def _autosave_enabled(self):
        try:
            return bool(self.prefs.get('autosave_enabled', True))
        except Exception:
            return True

    def _autosave_interval_min(self):
        try:
            v = int(self.prefs.get('autosave_interval', 5))
            return max(1, min(v, 60))
        except Exception:
            return 5

    def _do_autosave(self):
        """把当前项目保存到独立备份文件（不覆盖原文件），自动保存后仍标记未保存。"""
        try:
            if not getattr(self, '_dirty', False):
                return
            if not self._autosave_enabled():
                return
            if not hasattr(self, '_build_project_payload'):
                return
            path = self._autosave_backup_path()
            payload = self._build_project_payload()
            self._atomic_save(path, payload)
            try:
                from src.core import dbg
                dbg.log('autosave -> %s' % path)
            except Exception:
                pass
        except Exception:
            pass

    def _start_autosave(self):
        """启动自动保存定时器（间隔在首选项配置，默认 5 分钟）。"""
        try:
            if getattr(self, '_autosave_after', None) is not None:
                return

            def _tick():
                try:
                    self._do_autosave()
                except Exception:
                    pass
                try:
                    self._autosave_after = self.root.after(
                        self._autosave_interval_min() * 60 * 1000, _tick)
                except Exception:
                    pass

            self._autosave_after = self.root.after(
                self._autosave_interval_min() * 60 * 1000, _tick)
        except Exception:
            pass

    def _check_autosave_recovery(self):
        """启动时若存在自动保存备份，询问是否恢复。"""
        try:
            if not self._autosave_enabled():
                return
            path = self._autosave_backup_path()
            if not os.path.exists(path):
                return
            if not messagebox.askyesno(
                    '发现自动保存备份',
                    '检测到上次会话的自动保存备份:\n%s\n\n是否恢复？' % path,
                    parent=self.root):
                return
            # 静默恢复（不加入最近文件）
            self._open_project_file(path, from_recent=True, add_to_recent=False)
            self._mark_dirty()   # 备份内容仍需用户手动保存为主文件
        except Exception:
            pass

    def menu_save_project(self):
        path = filedialog.asksaveasfilename(
            title='保存项目', filetypes=[('SCS 项目文件', '*.scs')],
            defaultextension='.scs', parent=self.root)
        if not path:
            return
        try:
            payload = self._build_project_payload()
            back = self._atomic_save(path, payload)
            if back.get('version') not in (2, 3) or 'principal' not in back or 'waterlines' not in back:
                raise RuntimeError('保存后回读校验未通过，文件未覆盖。')
            self._clear_dirty()
            self._current_project_path = os.path.abspath(path)
            self._add_recent_project(path)
            self._refresh_statusbar()
            self.log('项目已成功保存到: %s' % path)
            # 确保 .scs 文件图标关联已注册（保存后刷新 Shell 图标缓存）
            try:
                self._register_scs_file_association()
            except Exception:
                pass
            messagebox.showinfo('保存成功', '项目已成功保存！\n%s' % path, parent=self.root)
        except Exception as e:
            messagebox.showerror('保存失败', '保存失败: %s' % e, parent=self.root)

    def menu_import_project(self, path=None):
        if path:
            self._open_project_file(path, from_recent=True)
            return
        path = filedialog.askopenfilename(
            title='选择要导入的项目文件', filetypes=[('SCS 项目文件', '*.scs'), ('所有文件', '*.*')],
            parent=self.root)
        if not path:
            return
        self._open_project_file(path)

    def _apply_project_logo(self, p):
        """从项目文件内嵌 logo 恢复窗口图标。"""
        logo = p.get('logo')
        if not logo:
            return
        import base64
        import io
        try:
            from PIL import Image, ImageTk
            img = Image.open(io.BytesIO(base64.b64decode(logo)))
            photo = ImageTk.PhotoImage(img)
            self.root.iconphoto(True, photo)
            self._project_logo_photo = photo  # 防止被 GC
        except Exception:
            pass

    def _apply_project_payload(self, p):
        """将项目数据恢复到当前应用（兼容旧版本文件：缺失字段用默认值）"""
        pr = p.get('principal', {})
        self.Lpp = pr.get('Lpp', math.nan)
        self.Breadth = pr.get('Breadth', math.nan)
        self.Depth = pr.get('Depth', math.nan)
        self.LppStartStation = pr.get('LppStartStation', math.nan)
        self.LppEndStation = pr.get('LppEndStation', math.nan)
        self.Draft = pr.get('Draft', math.nan)
        self.HeelAngle = pr.get('HeelAngle', 0)
        self.TrimAngle = pr.get('TrimAngle', 0)
        self.OriginFlag = pr.get('OriginFlag', 'amidship')
        self.CoefficientMethod = pr.get('CoefficientMethod', 'trapezoidal')
        self.waterlines = p.get('waterlines', [])
        self.decklines = p.get('decklines', [])
        self.bodyplans = p.get('bodyplans', [])
        self.sections = p.get('sections', {})
        self.Hydrostatics = p.get('hydrostatics')
        self.BonjeanCurves = p.get('bonjean')
        self.StabilityData = p.get('stability')
        self.GZ_CurveData = p.get('gz')
        self.DynamicStabilityData = p.get('dynamic')
        bd = p.get('buoyancy', {})
        self.BuoyancyVolume = bd.get('volume', math.nan)
        self.BuoyancyCenter = bd.get('center', [math.nan, math.nan, math.nan])
        self.SectionAreas = p.get('section_areas', self.SectionAreas)
        self.ML_model = p.get('ml_model')
        ui = p.get('ui_state', {})
        self.IsLocked = ui.get('IsLocked', False)
        self.IsSymmetricView = ui.get('IsSymmetricView', False)
        self.WireframeMode = ui.get('WireframeMode', '实体曲面')
        self.SurfaceColor = ui.get('SurfaceColor', '#ccccff')
        try:
            self.var_wiremode.set(self.WireframeMode)
        except Exception:
            pass
        # 表格与原始数据
        self.Half_table.set_columns(p.get('half_table_cols', self.Half_table.get_columns()))
        self.Half_table.set_data(p.get('half_table_data', []))
        self.Z_table.set_columns(p.get('z_table_cols', self.Z_table.get_columns()))
        self.Z_table.set_data(p.get('z_table_data', []))
        self.original_data = p.get('original_data', [])
        self.original_headers = p.get('original_headers', [])
        if self.original_headers:
            self.original_table.set_columns(self.original_headers)
            self.original_table.set_data(self.original_data)
        # 重建树
        self._rebuild_tree_from_data()
        self.update_half_width_plot()
        # 恢复各计算结果曲线（静水力/邦戎/稳性），保证"保存→重新打开"后仍可见
        self._restore_result_plots()
        self._update_button_state()
        self._refresh_statusbar()

    def _restore_result_plots(self):
        """项目加载后按已恢复的数据重绘静水力 / 邦戎 / 稳性曲线。"""
        try:
            hs = getattr(self, 'Hydrostatics', None)
            if hs and len(hs.get('drafts', [])) and getattr(self, '_plot_hydrostatics_curves', None):
                self._plot_hydrostatics_curves(hs)
        except Exception:
            pass
        try:
            br = getattr(self, 'BonjeanCurves', None)
            if br and getattr(self, '_plot_bonjean', None):
                self._plot_bonjean(br)
                self._fill_bonjean_table(br)
        except Exception:
            pass
        try:
            if getattr(self, 'StabilityData', None) and getattr(self, '_render_kn_plot', None):
                self._render_kn_plot()
        except Exception:
            pass
        try:
            if getattr(self, 'GZ_CurveData', None) and getattr(self, '_render_gz_plot', None):
                self._render_gz_plot()
        except Exception:
            pass
        try:
            dyn = getattr(self, 'DynamicStabilityData', None)
            gz = getattr(self, 'GZ_CurveData', None)
            if dyn and gz and getattr(self, '_render_dynamic_plot', None):
                heels = np.asarray(gz.get('HeelAngles', []), dtype=float)
                gzv = np.asarray(gz.get('GZ_Values', []), dtype=float)
                disp = gz.get('Displacement', 0)
                kg = gz.get('KG', 0)
                try:
                    from src.core import ship_core as _core
                    _, passed = _core.check_stability_regulations(
                        dyn, getattr(self, 'Breadth', None), getattr(self, 'Depth', None))
                except Exception:
                    passed = True
                self._render_dynamic_plot(dyn, heels, gzv, disp, kg, passed)
        except Exception:
            pass

    def _rebuild_tree_from_data(self):
        """按数据重建项目树：同一船型的水线面/横剖面归入同一"船型模型"节点"""
        for iid in self.tree.get_children(self.tree_root):
            self.tree.delete(iid)
            self.tree_meta.pop(iid, None)
        self.table_root = self.tree.insert(self.tree_root, 'end', text='Table', open=True)
        self.model_root = self.tree.insert(self.tree_root, 'end', text='Model', open=True)
        self.face_root = self.tree.insert(self.tree_root, 'end', text='Face', open=True)
        # 水线面：按船型分组
        ships = {}
        for wl in self.waterlines:
            key = wl.get('ship', '船型模型')
            ship = ships.get(key)
            if ship is None:
                ship = self._find_or_create_child(self.face_root, key,
                                                  {'type': 'ship_model_root'})
                ships[key] = ship
            half = self._find_or_create_child(ship, 'Half', {'type': 'half_root'})
            self._tree_add(half, wl['name'], {'type': 'waterline', 'table': wl['table'],
                                              'height': wl.get('height', math.nan)})
        # 横剖面：按船型分组
        for bp in self.bodyplans:
            key = bp.get('ship', '船型模型')
            ship = ships.get(key)
            if ship is None:
                ship = self._find_or_create_child(self.face_root, key,
                                                  {'type': 'ship_model_root'})
                ships[key] = ship
            body = self._find_or_create_child(ship, 'Body Plan', {'type': 'bodyplan_root'})
            self._tree_add(body, bp['name'], {'type': 'bodyplan', 'table': bp['table'],
                                              'station': bp['station']})
        # 已导入的原表格节点
        if self.original_headers:
            self._tree_add(self.table_root, '已导入表格', {
                'type': 'table', 'Data': self.original_data,
                'Headers': self.original_headers, 'VariableNames': self.original_headers})
        # 机器学习模型节点
        if self.ML_model:
            name = os.path.basename(self.ML_model.get('path') or '') or '已加载模型'
            self._tree_add(self.model_root, name,
                           {'type': 'model', 'path': self.ML_model.get('path')})

    # =====================================================================
    # 3D 视图（型线）
    # =====================================================================
    def menu_half_section(self):
        """三维型线视图"""
        if not self.waterlines and not self.bodyplans:
            messagebox.showinfo('无数据', '没有可绘制的型线数据，请先导入。', parent=self.root)
            return
        if not self._principal_ok():
            messagebox.showinfo('主尺度未设置', '请先设置主尺度。', parent=self.root)
            return
        ratio = self.Lpp / (self.LppEndStation - self.LppStartStation)
        dlg = tk.Toplevel(self.root)
        dlg.title('三维型线视图')
        dlg.geometry('700x560')
        canvas = PlotCanvas(dlg, three_d=True, toolbar=True)
        canvas.pack(fill='both', expand=True)
        ax = canvas.ax
        for wl in self.waterlines:
            table = wl.get('table', {})
            st = np.array([core._parse_scalar(v) for v in table.get('rows', [])])
            if st.size == 0:
                continue
            cols = {}
            for i, name in enumerate(table.get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in table['rows']]
            stations = np.array(cols.get('站号', []), dtype=float)
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = wl.get('height', math.nan)
            if not math.isfinite(z):
                continue
            z_data = np.full(stations.size, z)
            x = (stations - self.LppStartStation) * ratio
            o = np.argsort(x)
            ax.plot(x[o], hw[o], z_data[o], 'b-', linewidth=1)
            ax.plot(x[o], -hw[o], z_data[o], 'b-', linewidth=1)
        for bp in self.bodyplans:
            cols = {}
            for i, name in enumerate(bp['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in bp['table']['rows']]
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = np.array(cols.get('高度', []), dtype=float)
            x_pos = (bp['station'] - self.LppStartStation) * ratio
            ax.plot([x_pos] * len(z), hw, z, 'r-', linewidth=1)
            ax.plot([x_pos] * len(z), -hw, z, 'r-', linewidth=1)
        ax.set_xlabel('纵向位置 (m)')
        ax.set_ylabel('横向位置 (m)')
        ax.set_zlabel('高度 Z (m)')
        ax.set_title('船舶型线三维视图')
        ax.view_init(elev=25, azim=-135)
        canvas.refresh()
        ttk.Button(dlg, text='关闭', command=dlg.destroy).pack(pady=4)
