# -*- coding: utf-8 -*-
"""
ship_app_calc.py —— 高级计算（浮心 / 静水力 / 邦戎 / 稳性 / 3D 蒙皮）
"""

import os
import math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import numpy as np

from src.core import ship_core as core


def _with_busy(msg='正在计算...'):
    """长任务忙指示装饰器：进入置等待光标+状态栏提示，退出复位（try/finally）。"""
    def deco(fn):
        def wrapper(self, *args, **kwargs):
            try:
                self.set_busy(True, msg)
            except Exception:
                pass
            try:
                return fn(self, *args, **kwargs)
            finally:
                try:
                    self.set_busy(False)
                except Exception:
                    pass
        return wrapper
    return deco


def _fmt_cell(v):
    """将单元格值转为可写文本：np 标量转 python 标量；NaN/None 写空串。"""
    try:
        import numpy as _np
        if isinstance(v, _np.floating):
            v = float(v)
        elif isinstance(v, _np.integer):
            v = int(v)
        if hasattr(v, 'item'):
            v = v.item()
    except Exception:
        pass
    try:
        if v is None:
            return ''
        if isinstance(v, float) and not math.isfinite(v):
            return ''
        return v
    except Exception:
        return v


class ShipAppCalc:
    # =====================================================================
    # 辅助：收集水线面/横剖面数据
    # =====================================================================
    def _collect_waterlines(self):
        """将 waterlines/decklines 转为 calculate_waterplane_at_draft 需要的格式"""
        wls = []
        for wl in self.waterlines:
            cols = {}
            for i, name in enumerate(wl['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in wl['table']['rows']]
            if '站号' not in cols or '半宽' not in cols:
                continue
            stations = np.array(cols['站号'], dtype=float)
            half = np.array(cols['半宽'], dtype=float)
            height = wl.get('height', math.nan)
            if not math.isfinite(height):
                continue
            wls.append({'stations': stations, 'halfWidths': half, 'height': height})
        return wls

    def _collect_sections(self):
        """sections: dict station -> {'Y','Z'}"""
        return {st: {'Y': np.asarray(sec['Y'], dtype=float),
                     'Z': np.asarray(sec['Z'], dtype=float)}
                for st, sec in self.sections.items()}

    def _selected_ship_node(self):
        """从当前选中节点向上查找船型根节点"""
        sel = self._selected_node()
        if not sel:
            return None
        node = sel
        while node:
            meta = self.tree_meta.get(node, {})
            if meta.get('type') == 'ship_model_root':
                return node
            parent = self.tree.parent(node)
            if not parent:
                break
            node = parent
        return None

    # =====================================================================
    # 浮心计算
    # =====================================================================
    @_with_busy('正在计算浮心...')
    def buoyancy_calc_clicked(self):
        method = self.var_buoyancy_method.get()
        try:
            draft = float(self.var_draft.get())
        except Exception:
            draft = 1.0
        self.Draft = draft
        self.HeelAngle = float(self.var_heel.get())
        self.TrimAngle = float(self.var_trim.get())
        self.log('开始浮心计算...')
        if not self._check_principal():
            return
        if method.startswith('正浮态'):
            # 基于水线面
            wls = self._collect_waterlines()
            if not wls:
                messagebox.showerror('数据缺失', '没有水线面数据，请先从"识别结果"导入水线面。',
                                     parent=self.root)
                return
            try:
                res = core.calc_buoyancy_from_waterplane(wls, self.Lpp, self.LppStartStation,
                                                         self.LppEndStation, draft)
            except Exception as e:
                messagebox.showerror('计算失败', '计算水线面参数失败: %s' % e, parent=self.root)
                return
        else:
            # 基于横剖面
            if not self.sections:
                messagebox.showerror('数据缺失', '没有横剖面数据，请先导入 Body Plan。', parent=self.root)
                return
            station_values = np.array(sorted(self.sections.keys()), dtype=float)
            ratio = self.Lpp / (self.LppEndStation - self.LppStartStation)
            mid = (self.LppStartStation + self.LppEndStation) / 2
            x_coords = (station_values - mid) * ratio
            sections = [self.sections[st] for st in station_values]
            try:
                res = core.calc_buoyancy_from_sections(
                    sections, x_coords, self.HeelAngle, self.TrimAngle, draft,
                    self.Lpp, self.Breadth if math.isfinite(self.Breadth) else 10,
                    self.Depth if math.isfinite(self.Depth) else 10)
            except Exception as e:
                messagebox.showerror('计算失败', '计算浮心出错: %s' % e, parent=self.root)
                return
        self.BuoyancyVolume = res['volume']
        self.BuoyancyCenter = [res['xB'], res['yB'], res['zB']]
        self.var_vol.set('%.4f' % res['volume'])
        self.var_xB.set('%.4f' % res['xB'])
        self.var_yB.set('%.4f' % res['yB'])
        self.var_zB.set('%.4f' % res['zB'])
        self.log('浮心体积: %.4f m³' % res['volume'])
        self.log('浮心坐标: [%.4f, %.4f, %.4f] m' % (res['xB'], res['yB'], res['zB']))
        self._update_buoyancy_text(method, draft, res)
        self.plot_buoyancy_result(res)

    def _update_buoyancy_text(self, method, draft, res):
        lines = ['━━━ 计算结果 ━━━', '计算方法: %s' % method, '吃水: %.3f m' % draft, '',
                 '浮心体积: %.4f m³' % res['volume'],
                 '浮心X: %.4f m' % res['xB'],
                 '浮心Y: %.4f m' % res['yB'],
                 '浮心Z: %.4f m' % res['zB'], '', '计算完成！']
        self.TextArea_buoyancy.delete('1.0', 'end')
        self.TextArea_buoyancy.insert('1.0', '\n'.join(lines))

    def plot_buoyancy_result(self, res):
        ax = self.plot_buoyancy
        ax.clear()
        if not self.waterlines:
            ax.show_message('无水线面数据可绘制')
            return
        wls = self._collect_waterlines()
        try:
            wp = core.calculate_waterplane_at_draft(wls, self.Lpp, self.LppStartStation,
                                                    self.LppEndStation, self.Draft)
        except Exception:
            return
        x = (wp['stations'] - self.LppStartStation) * self.Lpp / (self.LppEndStation - self.LppStartStation)
        x = x - self.Lpp / 2
        o = np.argsort(x)
        x, y = x[o], wp['halfWidths'][o]
        ax.ax.plot(x, y, 'b-', linewidth=2.5, label='右舷水线面')
        ax.ax.plot(x, -y, 'b-', linewidth=2.5, label='左舷水线面')
        if x.size > 1:
            xf = np.concatenate([x, x[::-1]])
            yf = np.concatenate([y, -y[::-1]])
            ax.ax.fill(xf, yf, color='#80ccff', alpha=0.4)
        ax.ax.plot(res['xB'], res['yB'], 'ro', markersize=10, label='浮心投影')
        ax.ax.plot([res['xB'], res['xB']], [0, res['yB']], 'r--')
        ax.ax.plot([0, res['xB']], [res['yB'], res['yB']], 'r--')
        ax.ax.text(res['xB'] + 0.5, res['yB'] + 0.5,
                   '浮心\nX=%.3f\nY=%.3f\nZ=%.3f' % (res['xB'], res['yB'], res['zB']),
                   fontsize=9, color='darkred')
        ax.ax.set_xlabel('纵向位置 (m, 船中为0)')
        ax.ax.set_ylabel('横向位置 (m)')
        ax.ax.set_title('正浮态水线面 (吃水: %.3f m)' % self.Draft)
        ax.ax.grid(True, alpha=0.3)
        ax.ax.legend(loc='best', fontsize=8)
        ax.ax.axis('equal')
        ax.refresh()

    # =====================================================================
    # 静水力曲线
    # =====================================================================
    @_with_busy('正在计算静水力曲线...')
    def calc_curves_clicked(self):
        if not self._check_principal():
            return
        draft_min = float(self.var_draft_min.get())
        draft_max = float(self.var_draft_max.get())
        n_points = int(self.var_draft_steps.get())
        if draft_min >= draft_max:
            messagebox.showerror('参数错误', '最小吃水必须小于最大吃水！', parent=self.root)
            return
        if draft_min < 0:
            messagebox.showerror('参数错误', '吃水不能为负值！', parent=self.root)
            return
        wls = self._collect_waterlines()
        if not wls:
            messagebox.showerror('数据缺失', '没有水线面数据，请先导入水线面。', parent=self.root)
            return
        self.log('正在计算静水力曲线...')
        try:
            hs = core.calc_hydrostatics(
                wls, self.sections, self.Lpp, self.LppStartStation, self.LppEndStation,
                self.Breadth if math.isfinite(self.Breadth) else self.Lpp / 6,
                draft_min, draft_max, n_points,
                outlier_removal=bool(self.var_outlier.get()),
                outlier_threshold=float(self.var_outlier_th.get()))
        except Exception as e:
            self.TextArea_curves.insert('end', '计算失败！\n%s' % e)
            messagebox.showerror('计算错误', str(e), parent=self.root)
            return
        self.Hydrostatics = hs
        self._plot_hydrostatics_curves(hs)
        self.log('静水力曲线计算完成。')
        self.TextArea_curves.delete('1.0', 'end')
        self.TextArea_curves.insert('1.0', '计算完成！点数: %d' % n_points)

    def _plot_curve_y(self, canvas, val, drafts, val_name, title):
        canvas.clear()
        canvas.ax.plot(val, drafts, 'b-', linewidth=2)
        canvas.ax.plot(val, drafts, 'bo', markersize=4)
        canvas.ax.set_ylabel('吃水 T (m)')
        canvas.ax.set_xlabel(val_name)
        canvas.ax.set_title(title)
        canvas.ax.grid(True, alpha=0.3)
        canvas.refresh()

    def _plot_section_area_along_length(self, canvas):
        """横剖面面积曲线：X=站号/船长纵向位置，Y=该站型剖面全面积（与吃水无关）。

        与 MATLAB plotSectionAreaAlongLength 一致：读取 Body Plan 各站半宽型值，
        用分段线性 ∫y·dz 求半船面积再乘 2。
        """
        canvas.clear()
        stations = []
        areas = []
        # sections: dict station -> {'Y': [...], 'Z': [...]}（bodyplans 导入时建立）
        try:
            items = sorted(self.sections.items(),
                           key=lambda kv: float(kv[0]))
        except Exception:
            items = []
        for st, sec in items:
            try:
                half_y = np.asarray(sec.get('Y', []), dtype=float).ravel()
                z = np.asarray(sec.get('Z', []), dtype=float).ravel()
            except Exception:
                continue
            if half_y.size < 2 or z.size < 2:
                continue
            valid = np.isfinite(z) & np.isfinite(half_y)
            z = z[valid]
            half_y = half_y[valid]
            if z.size < 2:
                continue
            o = np.argsort(z)
            z = z[o]
            half_y = half_y[o]
            h = np.diff(z)
            half_area = float(np.sum(h * (half_y[:-1] + half_y[1:]) / 2.0))
            stations.append(float(st))
            areas.append(2.0 * half_area)  # 全船型剖面面积
        if not stations:
            canvas.show_message('无 Body Plan 横剖面数据，无法绘制横剖面面积曲线。\n'
                                '请先在"横剖面"页导入 Body Plan。',
                                '站号', '横剖面面积 A_s (m^2)')
            return
        # 横坐标使用站号（用户要求：横剖面曲线横坐标为站号）
        canvas.ax.plot(stations, areas, 'b-o', linewidth=1.5, markersize=4)
        canvas.ax.set_xlabel('站号')
        canvas.ax.set_ylabel('横剖面面积 A_s (m^2)')
        canvas.ax.set_title('横剖面面积曲线')
        canvas.ax.grid(True, alpha=0.3)
        canvas.refresh()

    def _plot_hydrostatics_curves(self, hs):
        d = hs['drafts']
        plots = self.curve_plots
        self._plot_curve_y(plots['水线面面积'], hs['Aw'], d, '水线面面积 Aw (m^2)', '水线面面积曲线')
        self._plot_section_area_along_length(plots['横剖面面积'])
        self._plot_curve_y(plots['排水量'], hs['dispMass'], d, '排水量 Δ (t)', '排水量曲线')
        self._plot_curve_y(plots['TPC'], hs['TPC'], d, 'TPC (t/cm)', 'TPC曲线')
        self._plot_curve_y(plots['MCT'], hs['MCT'], d, 'MCT (t·m/cm)', 'MCT曲线')
        self._plot_curve_y(plots['浮心X'], hs['LCB'], d, '浮心纵向位置 LCB (m)', '浮心纵向位置曲线')
        self._plot_curve_y(plots['漂心X'], hs['LCF'], d, '漂心纵向位置 LCF (m)', '漂心纵向位置曲线')
        self._plot_curve_y(plots['浮心Z'], hs['VCB'], d, '浮心垂向位置 VCB (m)', '浮心垂向位置曲线')
        # KM
        p = plots['稳心高度KM']
        p.clear()
        p.ax.plot(hs['KMT'], d, 'b-', linewidth=2, label='KMT')
        p.ax.plot(hs['KML'], d, 'r--', linewidth=2, label='KML')
        p.ax.set_ylabel('吃水 T (m)')
        p.ax.set_xlabel('高度 (m)')
        p.ax.set_title('稳心高度曲线')
        p.ax.legend(fontsize=8)
        p.ax.grid(True, alpha=0.3)
        p.refresh()
        # 系数
        p = plots['船型系数']
        p.clear()
        p.ax.plot(hs['CB'], d, 'b-', linewidth=2, label='Cb (方形)')
        p.ax.plot(hs['CP'], d, 'g-', linewidth=2, label='Cp (纵棱)')
        p.ax.plot(hs['CM'], d, 'r-', linewidth=2, label='Cm (中横)')
        p.ax.plot(hs['CW'], d, 'm-', linewidth=2, label='Cw (水线)')
        p.ax.set_ylabel('吃水 T (m)')
        p.ax.set_xlabel('系数')
        p.ax.set_title('船型系数曲线')
        p.ax.legend(fontsize=8)
        p.ax.grid(True, alpha=0.3)
        p.refresh()
        # 综合曲线
        p = plots['综合曲线']
        p.clear()
        p.ax.plot(hs['CB'] * 10, d, 'b-', linewidth=1.5, label='Cb x 10')
        p.ax.plot(hs['CP'] * 10, d, 'g-', linewidth=1.5, label='Cp x 10')
        p.ax.plot(hs['CM'] * 10, d, 'r-', linewidth=1.5, label='Cm x 10')
        p.ax.plot(hs['CW'] * 10, d, 'm-', linewidth=1.5, label='Cw x 10')
        max_disp = float(np.max(hs['dispMass'])) if len(hs['dispMass']) else 0
        if max_disp > 0:
            scale_disp = 10 ** math.floor(math.log10(max_disp / 5))
            p.ax.plot(hs['dispMass'] / scale_disp, d, 'k-', linewidth=2,
                      label='Disp (1cm=%.0ft)' % scale_disp)
        p.ax.plot(hs['LCB'], d, 'b:', label='LCB (m)')
        p.ax.plot(hs['LCF'], d, 'g:', label='LCF (m)')
        p.ax.plot(hs['VCB'], d, 'r:', label='VCB (m)')
        p.ax.set_ylabel('吃水 T (m)')
        p.ax.set_title('综合静水力曲线')
        p.ax.legend(loc='best', fontsize=7)
        p.ax.grid(True, alpha=0.3)
        p.refresh()

    # =====================================================================
    # 邦戎曲线
    # =====================================================================
    def _on_bonjean_all_changed(self):
        pass

    def _parse_station_list(self, s):
        """解析站号字符串：支持 0,1,2 / 0:10 / 空格分隔"""
        s = s.strip()
        out = []
        # 冒号区间
        for part in s.replace(',', ' ').split():
            if ':' in part:
                try:
                    segs = [float(x) for x in part.split(':')]
                    if len(segs) == 3:
                        out.extend(core.make_seq_prealloc(segs[0], segs[1], segs[2]))
                    elif len(segs) == 2:
                        out.extend(np.arange(segs[0], segs[1] + 1, 1.0))
                    else:
                        out.append(segs[0])
                except ValueError:
                    return None
            else:
                try:
                    out.append(float(part))
                except ValueError:
                    return None
        return np.array(out, dtype=float)

    @_with_busy('正在计算邦戎曲线...')
    def calc_bonjean_clicked(self):
        draft_min = float(self.var_bonjean_min.get())
        draft_max = float(self.var_bonjean_max.get())
        n_drafts = int(self.var_bonjean_steps.get())
        if draft_min >= draft_max:
            messagebox.showerror('参数错误', '最小吃水必须小于最大吃水！', parent=self.root)
            return
        if draft_min < 0:
            messagebox.showerror('参数错误', '吃水不能为负值！', parent=self.root)
            return
        if not self.sections:
            messagebox.showerror('数据缺失', '没有横剖面数据，请先导入 Body Plan。', parent=self.root)
            return
        if self.var_bonjean_all.get():
            stations = np.array(sorted(self.sections.keys()), dtype=float)
        else:
            stations = self._parse_station_list(self.var_bonjean_station.get())
            if stations is None or stations.size == 0:
                messagebox.showerror('参数错误', '站号格式不正确。', parent=self.root)
                return
            stations = np.sort(np.unique(stations))
        drafts = np.linspace(draft_min, draft_max, n_drafts)
        section_list = []
        station_positions = []
        for st in stations:
            if st in self.sections:
                section_list.append(self.sections[st])
                station_positions.append(self.get_station_position(st))
            else:
                section_list.append({'Y': [], 'Z': []})
                station_positions.append(math.nan)
        self.TextArea_bonjean.delete('1.0', 'end')
        self.TextArea_bonjean.insert('1.0', '正在计算邦戎曲线...')
        self.root.update_idletasks()
        res = core.calc_bonjean(section_list, station_positions, drafts)
        res['stations'] = stations
        self.BonjeanCurves = res
        self._plot_bonjean(res)
        self._fill_bonjean_table(res)
        self.TextArea_bonjean.delete('1.0', 'end')
        self.TextArea_bonjean.insert('1.0',
                                     '计算完成！\n站号数: %d\n吃水点数: %d' % (len(stations), n_drafts))
        self.log('邦戎曲线计算完成。')

    def get_station_position(self, station_num):
        if not math.isfinite(self.Lpp):
            return math.nan
        n = self.LppEndStation - self.LppStartStation
        if n == 0:
            return 0.0
        spacing = self.Lpp / n
        if self.OriginFlag == 'stern':
            return (station_num - self.LppStartStation) * spacing
        if self.OriginFlag == 'bow':
            return (self.LppEndStation - station_num) * spacing
        mid = (self.LppStartStation + self.LppEndStation) / 2
        return (station_num - mid) * spacing

    def _plot_bonjean(self, res):
        drafts = res['drafts']
        stations = res['stations']
        areas = res['areas']
        my = res['momentsY']
        mz = res['momentsZ']
        plots = self.bonjean_plots
        # 面积曲线
        p = plots['面积曲线']
        p.clear()
        colors = plt_cmap(len(stations))
        for i, st in enumerate(stations):
            p.ax.plot(areas[i], drafts, '-', color=colors[i % len(colors)],
                      label='站%s' % core.num2trimstr(st))
        p.ax.set_xlabel('面积 (m²)')
        p.ax.set_ylabel('吃水 (m)')
        p.ax.set_title('邦戎面积曲线')
        p.ax.legend(loc='best', fontsize=6)
        p.ax.grid(True, alpha=0.3)
        p.refresh()
        # 力矩
        p = plots['力矩Mz']
        p.clear()
        for i, st in enumerate(stations):
            p.ax.plot(mz[i], drafts, '-', color=colors[i % len(colors)],
                      label='站%s' % core.num2trimstr(st))
        p.ax.set_xlabel('面积矩 Mz (m³)')
        p.ax.set_ylabel('吃水 (m)')
        p.ax.set_title('邦戎面积矩曲线')
        p.ax.legend(loc='best', fontsize=6)
        p.ax.grid(True, alpha=0.3)
        p.refresh()
        # 形心Z
        p = plots['形心Z']
        p.clear()
        for i, st in enumerate(stations):
            cz = res['centroidsZ'][i]
            p.ax.plot(cz, drafts, '-', color=colors[i % len(colors)],
                      label='站%s' % core.num2trimstr(st))
        p.ax.set_xlabel('形心Z (m)')
        p.ax.set_ylabel('吃水 (m)')
        p.ax.set_title('邦戎形心曲线')
        p.ax.legend(loc='best', fontsize=6)
        p.ax.grid(True, alpha=0.3)
        p.refresh()
        # 综合图（模仿 MATLAB plotBonjeanComprehensiveView：
        # X = 站位纵向位置(船中为0)；面积/面积矩按站距归一化后在各站位两侧展开）
        p = plots['综合图']
        p.clear()
        positions = res.get('stationPositions')
        if positions is None or len(positions) != len(stations):
            positions = list(range(len(stations)))
        positions = np.asarray(positions, dtype=float)
        area_labeled = [False]
        mz_labeled = [False]
        for i, st in enumerate(stations):
            x = float(positions[i])
            As = areas[i]
            Mz = mz[i]
            valid = np.isfinite(As) & (As > 0)
            if np.sum(valid) < 2:
                continue
            # 站距（用于面积归一化，约展开到站距的 1/3）
            spacing = 1.0
            if i < len(positions) - 1:
                spacing = abs(float(positions[i + 1]) - x) or 1.0
            elif i > 0:
                spacing = abs(x - float(positions[i - 1])) or 1.0
            # 面积（实线）
            max_as = float(np.max(As[valid]))
            scale_as = spacing * 0.3 / max_as if max_as > 0 else 1.0
            lbl = '面积 As' if not area_labeled[0] else None
            p.ax.plot(x + As[valid] * scale_as, drafts[valid], '-',
                      linewidth=1.5, color='#0072BD', label=lbl)
            if lbl:
                area_labeled[0] = True
            # 面积矩（虚线）
            vm = np.isfinite(Mz) & valid
            if np.sum(vm) > 1:
                max_mz = float(np.max(np.abs(Mz[vm])))
                scale_mz = spacing * 0.3 / max_mz if max_mz > 0 else 1.0
                lbl = '面积矩 Mz' if not mz_labeled[0] else None
                p.ax.plot(x + Mz[vm] * scale_mz, drafts[vm], '--',
                          linewidth=1.2, color='#D95319', label=lbl)
                if lbl:
                    mz_labeled[0] = True
            # 站位中心线 + 站号标注
            p.ax.plot([x, x], [float(np.min(drafts[valid])), float(np.max(drafts[valid]))],
                      ':', linewidth=0.5, color='gray')
            p.ax.text(x, float(np.max(drafts[valid])) + 0.5,
                      '站%s' % core.num2trimstr(st), ha='center', fontsize=8)
        p.ax.set_xlabel('纵向位置 X (船中为0, m)')
        p.ax.set_ylabel('吃水 (m)')
        p.ax.set_title('邦戎曲线综合视图（面积实线 / 面积矩虚线）')
        if p.ax.get_legend_handles_labels()[0]:
            p.ax.legend(fontsize=8)
        p.ax.grid(True, alpha=0.3)
        if positions.size:
            xr = float(np.max(positions) - np.min(positions)) or 1.0
            p.ax.set_xlim(float(np.min(positions)) - xr * 0.1,
                          float(np.max(positions)) + xr * 0.1)
        p.refresh()

    def _fill_bonjean_table(self, res):
        rows = []
        for i, st in enumerate(res['stations']):
            for j, d in enumerate(res['drafts']):
                rows.append([core.num2trimstr(st), '%.3f' % d,
                             '%.4f' % res['areas'][i, j] if math.isfinite(res['areas'][i, j]) else '',
                             '%.4f' % res['centroidsY'][i, j] if math.isfinite(res['centroidsY'][i, j]) else '',
                             '%.4f' % res['centroidsZ'][i, j] if math.isfinite(res['centroidsZ'][i, j]) else ''])
        self.Bonjean_table.set_data(rows)

    # =====================================================================
    # 稳性计算
    # =====================================================================
    def _parse_range(self, s):
        """解析 MATLAB 冒号序列，如 0:10:90 或 1:5 或逗号列表。

        MATLAB 语义 a:b:c = a 起步长 b 到 c（含尾端点）：
          '0:10:90' -> [0,10,20,...,90]；'1:5' -> [1,2,3,4,5]（步长缺省为 1）。
        用整数步数 linspace 生成，避免浮点含尾端点丢失。
        """
        s = s.strip()
        if not s:
            return np.array([])
        if ':' in s:
            segs = [float(x) for x in s.split(':')]
            if len(segs) == 3:
                start, step, stop = segs[0], segs[1], segs[2]
                if step == 0:
                    return np.array([start])
                count = int(math.floor((stop - start) / step + 1e-9)) + 1
                if count < 1:
                    count = 1
                return np.linspace(start, start + step * (count - 1), count)
            if len(segs) == 2:
                start, stop = segs[0], segs[1]
                count = int(math.floor(stop - start + 1e-9)) + 1
                if count < 1:
                    count = 1
                return np.linspace(start, start + (count - 1), count)
            return np.array([segs[0]])
        return np.array([float(x) for x in s.replace(',', ' ').split()], dtype=float)

    def _buoyancy_fn(self):
        """构造浮心计算闭包供 KN 曲线使用"""
        def fn(heel, draft):
            old_draft = self.Draft
            old_heel = self.HeelAngle
            self.Draft = draft
            self.HeelAngle = heel
            try:
                if not self.sections:
                    raise ValueError('无横剖面数据')
                station_values = np.array(sorted(self.sections.keys()), dtype=float)
                ratio = self.Lpp / (self.LppEndStation - self.LppStartStation)
                mid = (self.LppStartStation + self.LppEndStation) / 2
                x_coords = (station_values - mid) * ratio
                sections = [self.sections[st] for st in station_values]
                res = core.calc_buoyancy_from_sections(
                    sections, x_coords, heel, 0, draft, self.Lpp,
                    self.Breadth if math.isfinite(self.Breadth) else 10,
                    self.Depth if math.isfinite(self.Depth) else 10)
                return res
            finally:
                self.Draft = old_draft
                self.HeelAngle = old_heel
        return fn

    @_with_busy('正在计算稳性横截曲线(KN)...')
    def calc_kn_clicked(self):
        if not self._check_principal():
            return
        if not self.sections:
            messagebox.showerror('数据缺失', '没有横剖面数据，请先导入 Body Plan。', parent=self.root)
            return
        heels = self._parse_range(self.var_stab_heels.get())
        drafts = self._parse_range(self.var_stab_drafts.get())
        if heels.size == 0 or drafts.size == 0:
            messagebox.showerror('参数错误', '横倾角或吃水范围格式错误。', parent=self.root)
            return
        self.log('开始计算稳性横截曲线(KN)...')
        self.TextArea_stability.delete('1.0', 'end')
        self.TextArea_stability.insert('1.0', '正在计算KN曲线...\n')
        self.root.update_idletasks()
        kn = core.calc_kn_curves(self._buoyancy_fn(), self.Lpp, self.LppStartStation,
                                 self.LppEndStation, self.Depth, heels, drafts)
        self.StabilityData = dict(KN_Curves=kn)
        self._render_kn_plot()
        self.TextArea_stability.delete('1.0', 'end')
        self.TextArea_stability.insert('1.0',
                                       'KN曲线计算完成！\n横倾角数: %d\n吃水点数: %d' % (len(heels), len(drafts)))
        self.log('KN曲线计算完成！')
        messagebox.showinfo('成功', 'KN曲线计算完成！', parent=self.root)

    def _render_kn_plot(self):
        """把 StabilityData['KN_Curves'] 绘制到 KN曲线 / 3D稳性曲面 两个子页。"""
        kn = (self.StabilityData or {}).get('KN_Curves')
        if not kn:
            return
        heels = np.asarray(kn['heels'], dtype=float).ravel()
        drafts = np.asarray(kn['drafts'], dtype=float).ravel()
        p = self.stability_plots['KN曲线']
        p.clear()
        colors = plt_cmap(max(len(heels), 1))
        for i in range(len(heels)):
            d_col = kn['Displacement'][:, i]
            k_col = kn['KN'][:, i]
            valid = np.isfinite(d_col) & np.isfinite(k_col)
            if np.sum(valid) > 1:
                p.ax.plot(d_col[valid], k_col[valid], '.-', color=colors[i % len(colors)],
                          linewidth=1.5, label='%.0f°' % heels[i])
        p.ax.set_xlabel('排水量 (t)')
        p.ax.set_ylabel('KN (m)')
        p.ax.set_title('稳性横截曲线 (KN Curves)')
        p.ax.grid(True, alpha=0.3)
        lines_with_labels = [l for l in p.ax.get_lines() if l.get_label()]
        if lines_with_labels:
            p.ax.legend(loc='best', fontsize=8)
        p.refresh()
        # 3D KR 曲面（Y=各吃水排水体积 / 1.025，与 MATLAB 一致）
        p = self.stability_plots['3D稳性曲面']
        p.clear()
        if len(heels) and len(drafts):
            X = np.repeat(heels[np.newaxis, :], len(drafts), axis=0)
            Y_vol = np.asarray(kn['Displacement'], dtype=float) / 1.025
            Z = np.asarray(kn['KN'], dtype=float)
            if np.sum(np.isfinite(Y_vol) & np.isfinite(Z)) >= 4:
                xx = X.copy()
                yy = Y_vol.copy()
                zz = Z.copy()
                yy[~np.isfinite(yy)] = math.nan
                zz[~np.isfinite(zz)] = math.nan
                p.ax.plot_surface(xx, yy, zz, alpha=0.7)
        p.ax.set_xlabel('横倾角 phi (deg)')
        p.ax.set_ylabel('排水体积 Vol (m^3)')
        p.ax.set_zlabel('形状稳性臂 KR (m)')
        p.ax.set_title('形状稳性曲面 KR(phi, V)')
        p.refresh()

    def _render_gz_plot(self):
        """把 GZ_CurveData 绘制到 GZ曲线 子页（保存恢复/查看用）。"""
        d = self.GZ_CurveData
        if d is None:
            return
        heels = np.asarray(d.get('HeelAngles', []), dtype=float).ravel()
        gz = np.asarray(d.get('GZ_Values', []), dtype=float).ravel()
        if heels.size == 0:
            return
        p = self.stability_plots['GZ曲线']
        p.clear()
        p.ax.plot(heels, gz, 'r-o', linewidth=2)
        p.ax.axhline(0, color='k', linewidth=1)
        p.ax.set_xlabel('横倾角 φ (deg)')
        p.ax.set_ylabel('复原力臂 GZ (m)')
        p.ax.set_title('静稳性曲线 (W=%.1ft, KG=%.2fm, YG=%.2fm)'
                      % (d.get('Displacement', 0), d.get('KG', 0), d.get('YG', 0)))
        p.ax.grid(True, alpha=0.3)
        p.refresh()

    @_with_busy('正在计算 GZ 曲线...')
    def calc_gz_clicked(self):
        if not self.StabilityData or 'KN_Curves' not in self.StabilityData:
            messagebox.showerror('数据缺失', '请先计算KN曲线！', parent=self.root)
            return
        kn = self.StabilityData['KN_Curves']
        # 默认值
        default_weight = 0.0
        default_kg = 0.0
        if math.isfinite(self.Depth):
            default_kg = 0.7 * self.Depth
            draft07 = 0.7 * self.Depth
            res = self._buoyancy_fn()(0, draft07)
            if math.isfinite(res['volume']):
                default_weight = res['volume'] * 1.025
        default_xg = 0.5 * self.Lpp if math.isfinite(self.Lpp) else 0
        from src.ui.ui_widgets import ask_numeric_dialog
        vals = ask_numeric_dialog(self.root, '输入船体重量和重心参数',
                                  ['船体重量 (t):', '重心X坐标 XG (m):',
                                   '重心Y坐标 YG (m):', '重心高度 KG (m):'],
                                  ['%.3f' % default_weight, '%.3f' % default_xg, '0.000', '%.3f' % default_kg])
        if vals is None:
            return
        ship_weight, xg, yg, kg = vals
        if ship_weight <= 0:
            messagebox.showerror('错误', '船体重量必须大于0。', parent=self.root)
            return
        gz = core.calc_gz_curve(kn, ship_weight, xg, yg, kg)
        self.GZ_CurveData = dict(HeelAngles=kn['heels'], GZ_Values=gz,
                                 Displacement=ship_weight, KG=kg, XG=xg, YG=yg)
        self._render_gz_plot()
        self.log('GZ曲线计算完成！排水量: %.3f t' % ship_weight)
        self.TextArea_stability.delete('1.0', 'end')
        self.TextArea_stability.insert('1.0', 'GZ曲线计算完成！\n排水量: %.3f t\n重心: KG=%.3f' % (ship_weight, kg))

    @_with_busy('正在计算动稳性...')
    def calc_dynamic_clicked(self):
        if self.GZ_CurveData is None:
            messagebox.showerror('数据缺失', '请先计算静稳性曲线(GZ)！', parent=self.root)
            return
        heels = np.asarray(self.GZ_CurveData['HeelAngles'], dtype=float)
        gz = np.asarray(self.GZ_CurveData['GZ_Values'], dtype=float)
        disp = self.GZ_CurveData['Displacement']
        kg = self.GZ_CurveData['KG']
        hs = self.Hydrostatics
        draft = self.Draft if math.isfinite(self.Draft) else 0
        dyn = core.calc_dynamic_stability(heels, gz, disp, kg, hydrostatics=hs,
                                          current_draft=draft)
        self.DynamicStabilityData = dyn
        # 法规检查
        check_results, is_passed = core.check_stability_regulations(dyn, self.Breadth, self.Depth)
        # 绘图（与 MATLAB 动稳性图标注一致）
        self._render_dynamic_plot(dyn, heels, gz, disp, kg, is_passed)
        # 稳性衡准数 K
        self.var_stab_k.set(dyn['stabilityKStatus'])
        # 报告
        lines = ['========== 稳性校核报告 ==========',
                 '依据: 《国内航行海船法定检验技术规则》(2011)', '',
                 '--- 校核项目 ---']
        lines.extend(check_results)
        lines += ['', '--- 计算参数 ---',
                  '排水量 Δ: %.1f t' % disp,
                  '重心高度 KG: %.2f m' % kg,
                  '初稳性高 GM: %.3f m' % dyn['GM'],
                  '最大复原力臂 GZ_max: %.3f m @ %.1f°' % (dyn['maxGZ'], dyn['angleMaxGZ']),
                  '稳性消失角: %.1f°' % dyn['vanishAngle'],
                  '最小倾覆力臂 l_q: %.4f m @ %.1f° (极限动倾角)' % (dyn['lq'], dyn['lq_angle']),
                  '最大风倾力臂 l_f: %.4f m (等面积法)' % dyn['lf'],
                  '', '==================================',
                  '总体结论: 【%s】' % ('合格' if is_passed else '不合格')]
        self.TextArea_stability.delete('1.0', 'end')
        self.TextArea_stability.insert('1.0', '\n'.join(lines))
        self.log('稳性计算完成。')
        messagebox.showinfo('完成', '动稳性计算完成！', parent=self.root)

    def _render_dynamic_plot(self, dyn, heels, gz, disp, kg, is_passed=True):
        """动稳性图（含 OFG/GHK 等面积、GM/lq 切线等 MATLAB 标注）。

        计算回调与"加载项目后恢复显示"共用。
        """
        p = self.stability_plots['动稳性']
        p.clear()
        ax1 = p.ax                       # 左轴：动稳性力臂 l_d
        ax2 = ax1.twinx()                # 右轴：静稳性力臂 GZ
        max_gz = float(dyn['maxGZ'])
        ang_max = float(dyn['angleMaxGZ'])
        ang_van = float(dyn['vanishAngle'])
        gm = float(dyn['GM']) if math.isfinite(dyn['GM']) else math.nan
        lq = float(dyn['lq']) if math.isfinite(dyn['lq']) else math.nan
        lf = float(dyn['lf']) if math.isfinite(dyn['lf']) else math.nan
        tg = float(dyn['theta_G']) if math.isfinite(dyn['theta_G']) else math.nan
        tk = float(dyn['theta_K']) if math.isfinite(dyn['theta_K']) else math.nan
        la_idx = int(dyn['lq_angle_idx'])
        # 左轴动稳性曲线
        ax1.plot(heels, dyn['dynamicArm'], 'b-o', linewidth=2, label='动稳性力臂 l_d')
        ax1.set_xlabel('横倾角 φ (deg)')
        ax1.set_ylabel('动稳性力臂 l_d (m·rad)')
        ax1.axhline(0, color='k', linewidth=0.8)
        # 右轴 GZ 曲线
        ax2.plot(heels, gz, 'r--', linewidth=1.5, label='静稳性力臂 GZ')
        ax2.set_ylabel('静稳性力臂 GZ (m)')
        # ---- GZ 特征点（右轴）----
        ax2.plot(ang_max, max_gz, 'r.', markersize=15)
        ax2.annotate('GZ_max=%.3fm\n@ %.1f°(极限静倾角)' % (max_gz, ang_max),
                     xy=(ang_max, max_gz), xytext=(ang_max + 1, max_gz * 1.05),
                     color='r', fontsize=9, va='bottom')
        ax2.plot(ang_van, 0, 'rx', markersize=10, linewidth=2)
        ax2.annotate('消失角 %.1f°' % ang_van, xy=(ang_van, 0),
                     xytext=(ang_van, -0.12 * (abs(max_gz) if abs(max_gz) > 1e-9 else 1.0)),
                     color='r', fontsize=9, ha='center', va='top')
        # ---- 等面积法示意（右轴，OFG/GHK、F/G/K 点、水平线 FK）----
        if math.isfinite(lf) and lf > 0 and math.isfinite(tg) and math.isfinite(tk):
            ax2.plot([0, tk], [lf, lf], 'm-', linewidth=2,
                     label='水平线FK (l_f=%.3fm)' % lf)
            ax2.plot(0, lf, 'ms', markersize=9, markerfacecolor='m')          # F
            ax2.plot(tg, lf, 'mo', markersize=9, markerfacecolor='y')          # G
            ax2.plot(tk, lf, 'md', markersize=9, markerfacecolor='c')          # K
            ax2.plot([tg, tg], [0, lf], 'm:', linewidth=1)
            ax2.fill([0, 0, tg, tg], [0, lf, lf, 0], color=(1, 0.8, 0.8),
                     alpha=0.3, linewidth=0)                                    # OFG
            ax2.text(tg / 2, lf / 2, 'OFG', color=(0.8, 0, 0), fontsize=10,
                     fontweight='bold', ha='center')
            idx_gk = np.nonzero((heels >= tg) & (heels <= tk))[0]
            if idx_gk.size >= 1:
                fx = np.concatenate([[tg], heels[idx_gk], [tk]])
                fy = np.concatenate([[lf], gz[idx_gk], [lf]])
                ax2.fill(fx, fy, color=(0.8, 1, 0.8), alpha=0.3, linewidth=0)   # GHK
                ax2.text((tg + tk) / 2, lf + (max_gz - lf) / 3, 'GHK',
                         color=(0, 0.6, 0), fontsize=10, fontweight='bold',
                         ha='center')
        # ---- GM 切线（57.3° 法则，右轴）----
        if math.isfinite(gm) and gm > 0:
            tmax = max(max_gz * 1.5, gm * 1.1)
            tx = [0.0, 57.3]
            ty = [0.0, gm]
            if gm > tmax and tmax > 1e-9:          # 超出纵轴时截断
                s = tmax / gm
                tx = [0.0, 57.3 * s]
                ty = [0.0, tmax]
            ax2.plot(tx, ty, 'k-.', linewidth=0.8, label='GM切线')
            ax2.plot([57.3, 57.3], [0, ty[-1]], 'k:', linewidth=0.5)
            ax2.plot(57.3, ty[-1], 'k*', markersize=8)
            ax2.text(57.3, ty[-1], '  GM=%.2fm' % gm, color='k', fontsize=9, va='bottom')
        # ---- 动稳性曲线特征（左轴）----
        try:
            ld_van = float(np.interp(ang_van, heels, dyn['dynamicArm']))
        except Exception:
            ld_van = math.nan
        if math.isfinite(ld_van):
            ax1.plot(ang_van, ld_van, 'b^', markersize=8)
            ax1.text(ang_van, ld_van, '  l_d=%.3f' % ld_van, color='b',
                     fontsize=9, va='bottom')
        if math.isfinite(lq) and lq > 0 and 0 < la_idx < len(dyn['dynamicArm']):
            la_ang = float(heels[la_idx])
            la_val = float(dyn['dynamicArm'][la_idx])
            ax1.plot([0, la_ang], [0, la_val], 'g-', linewidth=1.5,
                     label='最小倾覆力臂切线 l_q=%.4fm' % lq)
            ax1.plot(la_ang, la_val, 'go', markersize=8, markerfacecolor='g')
            ax1.text(la_ang, la_val, '  l_q=%.4fm\n  @ %.1f°' % (lq, la_ang),
                     color=(0, 0.5, 0), fontsize=9, va='bottom')
            end_deg = max(float(np.max(heels)), 57.3)
            ax1.plot([la_ang, end_deg], [la_val, lq * math.radians(end_deg)],
                     'g--', linewidth=1)
            ax1.plot(end_deg, lq * math.radians(end_deg), 'g*', markersize=9)
            ax1.text(end_deg, lq * math.radians(end_deg), '  l_q=%.4fm\n  (57.3°处)' % lq,
                     color=(0, 0.5, 0), fontsize=9, va='bottom')
        h1, l1 = ax1.get_legend_handles_labels()
        h2, l2 = ax2.get_legend_handles_labels()
        ax1.legend(h1 + h2, l1 + l2, loc='best', fontsize=8)
        ax1.grid(True, alpha=0.3)
        try:
            status_txt = '【%s】' % ('合格' if is_passed else '不合格')
            ax1.set_title('动稳性曲线 (Δ=%.1ft, KG=%.2fm, GM=%.2fm) %s'
                          % (disp, kg, gm if math.isfinite(gm) else 0.0, status_txt))
        except Exception:
            pass
        p.refresh()

    def export_stability(self):
        """导出动稳性曲线数据（txt/csv/xlsx），含完整校核报告参数。"""
        if self.DynamicStabilityData is None:
            messagebox.showerror('无数据', '请先计算动稳性曲线！', parent=self.root)
            return
        path = filedialog.asksaveasfilename(
            title='导出动稳性数据',
            filetypes=[('Excel 工作簿', '*.xlsx'), ('CSV 文件', '*.csv'),
                       ('文本文件', '*.txt')],
            defaultextension='.xlsx', parent=self.root)
        if not path:
            return
        try:
            d = self.DynamicStabilityData
            heels = np.asarray(d['heels'], dtype=float).ravel()
            gz = np.asarray(d['GZ'], dtype=float).ravel()
            darm = np.asarray(d['dynamicArm'], dtype=float).ravel()
            headers = ['HeelAngle_deg', 'GZ_m', 'DynamicArm_m_rad']
            rows = [[heels[i], gz[i], darm[i]] for i in range(len(heels))]
            import datetime
            meta_rows = [['Displacement_t', d.get('Displacement', self.GZ_CurveData.get('Displacement') if self.GZ_CurveData else '')],
                         ['KG_m', d.get('KG', '')],
                         ['GM_m', d.get('GM', '')],
                         ['StabilityK', d.get('stabilityK', '')],
                         ['CalculationTime', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')]]
            self._save_tabular(path, headers, rows,
                               extra_sheets={'计算参数': (['Parameter', 'Value'], meta_rows)})
            messagebox.showinfo('导出成功', '数据已导出。', parent=self.root)
        except Exception as e:
            messagebox.showerror('导出失败', str(e), parent=self.root)

    def _save_tabular(self, path, headers, rows, extra_sheets=None):
        """通用保存：.txt/.csv 逗号分隔；.xlsx 用 openpyxl（可含多 sheet）。"""
        low = (path or '').lower()
        if low.endswith('.xlsx'):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Data'
            ws.append(headers)
            for r in rows:
                ws.append([_fmt_cell(c) for c in r])
            for sheet_name, (h2, r2) in (extra_sheets or {}).items():
                ws2 = wb.create_sheet(sheet_name)
                ws2.append(h2)
                for r in r2:
                    ws2.append([_fmt_cell(c) for c in r])
            wb.save(path)
            return
        with open(path, 'w', encoding='utf-8', newline='') as f:
            import csv as _csv
            w = _csv.writer(f, lineterminator='\n')
            w.writerow(headers)
            for r in rows:
                w.writerow([_fmt_cell(c) for c in r])

    def export_kn(self):
        """导出 KN 稳性横截曲线数据（MATLAB ExportStabilityButtonPushed 对应）。
        行=吃水，列=横倾角，首列为 Draft_m。"""
        if not self.StabilityData or 'KN_Curves' not in self.StabilityData:
            messagebox.showerror('数据缺失', '没有可导出的稳性数据，请先计算KN曲线！', parent=self.root)
            return
        kn = self.StabilityData['KN_Curves']
        path = filedialog.asksaveasfilename(
            title='导出KN数据',
            filetypes=[('Excel 工作簿', '*.xlsx'), ('CSV 文件', '*.csv'),
                       ('文本文件', '*.txt')],
            defaultextension='.xlsx', parent=self.root)
        if not path:
            return
        try:
            heels = np.asarray(kn['heels'], dtype=float).ravel()
            drafts = np.asarray(kn['drafts'], dtype=float).ravel()
            kmat = np.asarray(kn['KN'], dtype=float)
            headers = ['Draft_m'] + ['Heel_%d_deg' % int(round(h)) for h in heels]
            rows = []
            for j, t in enumerate(drafts):
                rows.append([t] + list(kmat[j, :]))
            self._save_tabular(path, headers, rows)
            self.log('KN数据已导出：%s' % path)
            messagebox.showinfo('导出成功', 'KN数据导出成功！', parent=self.root)
        except Exception as e:
            messagebox.showerror('导出失败', str(e), parent=self.root)

    def export_gz(self):
        """导出 GZ 静稳性曲线数据（MATLAB ExportGZCurveButtonPushed 对应）。"""
        if self.GZ_CurveData is None or not len(self.GZ_CurveData.get('GZ_Values', []) or []):
            messagebox.showerror('数据缺失', '没有可导出的GZ曲线数据，请先计算GZ曲线！', parent=self.root)
            return
        path = filedialog.asksaveasfilename(
            title='导出GZ曲线数据',
            filetypes=[('Excel 工作簿', '*.xlsx'), ('CSV 文件', '*.csv'),
                       ('文本文件', '*.txt')],
            defaultextension='.xlsx', parent=self.root)
        if not path:
            return
        try:
            d = self.GZ_CurveData
            heels = np.asarray(d['HeelAngles'], dtype=float).ravel()
            gz = np.asarray(d['GZ_Values'], dtype=float).ravel()
            headers = ['HeelAngle_deg', 'GZ_m']
            rows = [[heels[i], gz[i]] for i in range(len(heels))]
            import datetime
            meta_rows = [['Displacement_t', d.get('Displacement', '')],
                         ['KG_m', d.get('KG', '')],
                         ['CalculationTime', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')]]
            self._save_tabular(path, headers, rows,
                               extra_sheets={'Metadata': (['Parameter', 'Value'], meta_rows)})
            self.log('GZ数据已导出：%s' % path)
            messagebox.showinfo('导出成功', 'GZ曲线数据导出成功！', parent=self.root)
        except Exception as e:
            messagebox.showerror('导出失败', str(e), parent=self.root)

    # =====================================================================
    # 3D 曲面生成
    # =====================================================================
    def _3d_principal(self):
        if not self._principal_ok():
            messagebox.showinfo('主尺度未设置', '请先设置主尺度。', parent=self.root)
            return None
        return (self.Lpp, self.LppStartStation, self.LppEndStation)

    @_with_busy('正在生成点云...')
    def gen_pointcloud_clicked(self):
        pr = self._3d_principal()
        if pr is None:
            return
        lpp, s0, s1 = pr
        ratio = lpp / (s1 - s0)
        xs, ys, zs, stations = [], [], [], []
        for wl in self.waterlines:
            cols = {}
            for i, name in enumerate(wl['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in wl['table']['rows']]
            st = np.array(cols.get('站号', []), dtype=float)
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = wl.get('height', math.nan)
            if not math.isfinite(z):
                continue
            valid = np.isfinite(st) & np.isfinite(hw)
            xs.extend(((st[valid] - s0) * ratio).tolist())
            ys.extend(hw[valid].tolist())
            zs.extend([z] * int(np.sum(valid)))
            stations.extend(st[valid].tolist())
        for bp in self.bodyplans:
            cols = {}
            for i, name in enumerate(bp['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in bp['table']['rows']]
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = np.array(cols.get('高度', []), dtype=float)
            valid = np.isfinite(hw) & np.isfinite(z)
            x_pos = (bp['station'] - s0) * ratio
            xs.extend([x_pos] * int(np.sum(valid)))
            ys.extend(hw[valid].tolist())
            zs.extend(z[valid].tolist())
            stations.extend([bp['station']] * int(np.sum(valid)))
        if not xs:
            messagebox.showerror('无数据', '在选定船型中未能收集到任何有效的三维点数据。', parent=self.root)
            return
        half_pts = np.column_stack([xs, ys, zs])
        mirrored = half_pts.copy()
        mirrored[:, 1] = -mirrored[:, 1]
        full_pts = np.unique(np.vstack([half_pts, mirrored]), axis=0)
        self.SurfaceGenerationData = dict(HalfPoints=half_pts, AllPoints=full_pts,
                                          HalfStations=np.array(stations))
        self._redraw_pointcloud()
        self._push_qt_pointcloud()
        self.log('点云生成成功，共 %d 个点。' % full_pts.shape[0])

    def _redraw_pointcloud(self, bottom_pts=None):
        """重新绘制 3D 点云视图。

        与 MATLAB 一致：原始点为红色，补齐底部的新增点为蓝色。
        点云降采样 + plot 标记（比 scatter 在 3D 中快得多）。
        """
        p = self.plot_face_area
        p.clear()
        all_pts = self.SurfaceGenerationData.get('AllPoints', np.empty((0, 3)))
        pts = all_pts
        if pts.shape[0] > 6000:
            sel = np.linspace(0, pts.shape[0] - 1, 6000).astype(int)
            pts = pts[sel]
        if len(pts):
            p.ax.plot(pts[:, 0], pts[:, 1], pts[:, 2], 'r.', markersize=1.5, ls='none')
        if bottom_pts is not None and len(bottom_pts):
            bm = np.vstack([bottom_pts, bottom_pts.copy() * np.array([1, -1, 1])])
            p.ax.plot(bm[:, 0], bm[:, 1], bm[:, 2], 'b.', markersize=2.5, ls='none')
        p.ax.set_xlabel('X (m)')
        p.ax.set_ylabel('Y (m)')
        p.ax.set_zlabel('Z (m)')
        p.ax.set_title('点云（红色：原始，蓝色：补齐底部）' if bottom_pts is not None and len(bottom_pts)
                       else '点云')
        p.ax.view_init(elev=25, azim=-135)
        p.set_true_box_aspect()
        p.refresh()

    @staticmethod
    def _polyline_to_segments(pts):
        """把 (N,3) 折线点序列转为 (N-1,2,3) 线段数组；点不足 2 返回 None。"""
        if pts is None or len(pts) < 2:
            return None
        pts = np.asarray(pts, dtype=float)
        return np.stack([pts[:-1], pts[1:]], axis=1)

    def _push_qt_pointcloud(self):
        """把点云推给 Qt 3D 视窗显示（无 Qt 时静默）。"""
        if self.qt3d_host is None:
            return
        pts = self.SurfaceGenerationData.get('AllPoints')
        if pts is None or len(pts) == 0:
            return
        pts = np.asarray(pts, dtype=float)
        if pts.shape[0] > 20000:
            sel = np.linspace(0, pts.shape[0] - 1, 20000).astype(int)
            pts = pts[sel]
        self.qt3d_host.set_pointcloud(pts, (1.0, 0.15, 0.15), 2.0)

    def _push_qt_lines(self):
        """把型线（水线/横剖面/底部轮廓）推给 Qt 3D 视窗（无 Qt 时静默）。"""
        if self.qt3d_host is None:
            return
        pr = self._3d_principal()
        if pr is None:
            return
        lpp, s0, s1 = pr
        ratio = lpp / (s1 - s0)
        groups = []
        # 水线：蓝色
        for wl in self.waterlines:
            cols = {}
            for i, name in enumerate(wl['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in wl['table']['rows']]
            st = np.array(cols.get('站号', []), dtype=float)
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = wl.get('height', math.nan)
            if not math.isfinite(z):
                continue
            valid = np.isfinite(st) & np.isfinite(hw)
            st, hw = st[valid], hw[valid]
            if st.size < 2:
                continue
            x = (st - s0) * ratio
            o = np.argsort(x)
            x, hw = x[o], hw[o]
            pos = self._polyline_to_segments(np.column_stack([x, hw, np.full(st.size, z)]))
            neg = self._polyline_to_segments(np.column_stack([x, -hw, np.full(st.size, z)]))
            if pos is not None:
                groups.append((pos, (0.25, 0.55, 0.95), 1.2))
            if neg is not None:
                groups.append((neg, (0.25, 0.55, 0.95), 1.2))
        # 横剖面：红色
        for bp in self.bodyplans:
            cols = {}
            for i, name in enumerate(bp['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in bp['table']['rows']]
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = np.array(cols.get('高度', []), dtype=float)
            valid = np.isfinite(hw) & np.isfinite(z)
            hw, z = hw[valid], z[valid]
            if z.size < 2:
                continue
            x_pos = (bp['station'] - s0) * ratio
            pos = self._polyline_to_segments(np.column_stack([np.full(z.size, x_pos), hw, z]))
            neg = self._polyline_to_segments(np.column_stack([np.full(z.size, x_pos), -hw, z]))
            if pos is not None:
                groups.append((pos, (0.85, 0.25, 0.25), 1.2))
            if neg is not None:
                groups.append((neg, (0.85, 0.25, 0.25), 1.2))
        # 底部轮廓与龙骨线：绿色
        bps = self.SurfaceGenerationData.get('BottomPoints')
        if bps is not None and len(bps):
            bps = np.asarray(bps, dtype=float)
            for xb in np.unique(bps[:, 0]):
                seg = bps[np.abs(bps[:, 0] - xb) < 1e-6]
                o = np.argsort(seg[:, 1])
                seg = seg[o]
                seg_pos = self._polyline_to_segments(seg[:, :3])
                if seg_pos is not None:
                    groups.append((seg_pos, (0.25, 0.75, 0.30), 1.0))
                    groups.append((seg_pos * np.array([1, -1, 1]), (0.25, 0.75, 0.30), 1.0))
            keel = bps[np.abs(bps[:, 1]) < 1e-6]
            if len(keel) >= 2:
                o = np.argsort(keel[:, 0])
                keel_seg = self._polyline_to_segments(keel[o, :3])
                if keel_seg is not None:
                    groups.append((keel_seg, (0.25, 0.75, 0.30), 1.5))
        self.qt3d_host.set_lines(groups)

    @_with_busy('正在绘制型线...')
    def gen_lines_clicked(self):
        pr = self._3d_principal()
        if pr is None:
            return
        lpp, s0, s1 = pr
        ratio = lpp / (s1 - s0)
        p = self.plot_face_area
        p.clear()
        for wl in self.waterlines:
            cols = {}
            for i, name in enumerate(wl['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in wl['table']['rows']]
            st = np.array(cols.get('站号', []), dtype=float)
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = wl.get('height', math.nan)
            valid = np.isfinite(st) & np.isfinite(hw)
            st, hw = st[valid], hw[valid]
            if st.size < 2:
                continue
            x = (st - s0) * ratio
            o = np.argsort(x)
            z_data = np.full(st.size, z)
            p.ax.plot(x[o], hw[o], z_data[o], 'b-', linewidth=1)
            p.ax.plot(x[o], -hw[o], z_data[o], 'b-', linewidth=1)
        for bp in self.bodyplans:
            cols = {}
            for i, name in enumerate(bp['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in bp['table']['rows']]
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = np.array(cols.get('高度', []), dtype=float)
            valid = np.isfinite(hw) & np.isfinite(z)
            hw, z = hw[valid], z[valid]
            if z.size < 2:
                continue
            x_pos = (bp['station'] - s0) * ratio
            p.ax.plot([x_pos] * z.size, hw, z, 'r-', linewidth=1)
            p.ax.plot([x_pos] * z.size, -hw, z, 'r-', linewidth=1)
        # 底部轮廓与龙骨线（补齐底部后显示，使型线图含船底，仿 MATLAB 型线图）
        bps = self.SurfaceGenerationData.get('BottomPoints')
        if bps is not None and len(bps):
            bps = np.asarray(bps, dtype=float)
            for xb in np.unique(bps[:, 0]):
                seg = bps[np.abs(bps[:, 0] - xb) < 1e-6]
                o = np.argsort(seg[:, 1])
                seg = seg[o]
                p.ax.plot(seg[:, 0], seg[:, 1], seg[:, 2], 'g-', linewidth=0.8)
                p.ax.plot(seg[:, 0], -seg[:, 1], seg[:, 2], 'g-', linewidth=0.8)
            keel = bps[np.abs(bps[:, 1]) < 1e-6]
            if len(keel) >= 2:
                o = np.argsort(keel[:, 0])
                p.ax.plot(keel[o, 0], keel[o, 1], keel[o, 2], 'g-', linewidth=1.5)
                p.ax.plot(keel[o, 0], keel[o, 1], keel[o, 2], 'g.', markersize=3)
        p.ax.set_xlabel('X (m)')
        p.ax.set_ylabel('Y (m)')
        p.ax.set_zlabel('Z (m)')
        p.ax.set_title('型线')
        p.ax.view_init(elev=25, azim=-135)
        p.set_true_box_aspect()
        p.refresh()
        self._push_qt_lines()
        self.log('船体型线已成功绘制。')

    @staticmethod
    def _densify_curve(x, y, n):
        """把型线 (x, y) 平滑加密为 n 个点（仿 MATLAB pchip 加密）。

        - x/y 允许乱序/重复：先按 x 排序去重；
        - 采用 PCHIP 保形插值避免线性加密的折角；
        - 数量不足 2 时返回空数组；半宽类数据裁剪到 >=0。
        返回 (xs, ys) 两个 ndarray。
        """
        x = np.asarray(x, dtype=float)
        y = np.asarray(y, dtype=float)
        m = np.isfinite(x) & np.isfinite(y)
        x, y = x[m], y[m]
        if x.size < 2:
            return np.array([]), np.array([])
        ux, idx = np.unique(x, return_index=True)
        uy = y[idx]
        if ux.size < 2:
            return np.array([]), np.array([])
        try:
            from scipy.interpolate import PchipInterpolator
            pp = PchipInterpolator(ux, uy, extrapolate=True)
            xs = np.linspace(ux[0], ux[-1], max(2, int(n)))
            ys = pp(xs)
        except Exception:
            xs = np.linspace(ux[0], ux[-1], max(2, int(n)))
            ys = np.interp(xs, ux, uy)
        return xs, np.clip(ys, 0.0, None)

    @_with_busy('正在生成船体蒙皮...')
    def gen_hull_clicked(self):
        """基于型线插值生成船体蒙皮（含甲板线顶部约束，仿 MATLAB natural 曲面）"""
        pr = self._3d_principal()
        if pr is None:
            return
        lpp, s0, s1 = pr
        ratio = lpp / (s1 - s0)
        # 收集并加密点云（含甲板线：顶部随甲板边线起伏，而非一刀切）
        dense = []
        # 水线：固定高度 z，半宽随站号 pchip 加密
        for wl in self.waterlines:
            cols = {}
            for i, name in enumerate(wl['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in wl['table']['rows']]
            st = np.array(cols.get('站号', []), dtype=float)
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = wl.get('height', math.nan)
            if not math.isfinite(z):
                continue
            st_f, hw_f = self._densify_curve(st, hw, 400)
            if st_f.size < 2:
                continue
            x_f = (st_f - s0) * ratio
            dense.append(np.column_stack([x_f, hw_f, np.full_like(st_f, z)]))
        # 甲板线（deckline）：半宽与高度都随站号变化，加密后提供顶部边界
        for dl in getattr(self, 'decklines', []) or []:
            cols = {}
            for i, name in enumerate(dl['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in dl['table']['rows']]
            st = np.array(cols.get('站号', []), dtype=float)
            hw = np.array(cols.get('半宽', []), dtype=float)
            zd = np.array(cols.get('高度', []) if '高度' in cols else [], dtype=float)
            if st.size < 2 or hw.size != st.size:
                continue
            if zd.size == st.size:
                # 两舷逐站高度：pchip 加密 x→(y,z)
                st_f, hw_f = self._densify_curve(st, hw, 400)
                _, zd_f = self._densify_curve(st, zd, 400)
                if st_f.size < 2:
                    continue
                x_f = (st_f - s0) * ratio
                dense.append(np.column_stack([x_f, hw_f, zd_f]))
            else:
                # 无逐站高度：按固定高度处理（兼容旧格式）
                z = dl.get('height', math.nan)
                if not math.isfinite(z):
                    continue
                st_f, hw_f = self._densify_curve(st, hw, 400)
                if st_f.size < 2:
                    continue
                x_f = (st_f - s0) * ratio
                dense.append(np.column_stack([x_f, hw_f, np.full_like(st_f, z)]))
        # 横剖面：半宽随高度 pchip 加密
        for bp in self.bodyplans:
            cols = {}
            for i, name in enumerate(bp['table'].get('columns', [])):
                cols[name] = [core._parse_scalar(r[i]) for r in bp['table']['rows']]
            hw = np.array(cols.get('半宽', []), dtype=float)
            z = np.array(cols.get('高度', []), dtype=float)
            valid = np.isfinite(hw) & np.isfinite(z)
            hw, z = hw[valid], z[valid]
            z_f, hw_f = self._densify_curve(z, hw, 200)
            if z_f.size < 2:
                continue
            x_pos = (bp['station'] - s0) * ratio
            dense.append(np.column_stack([np.full_like(z_f, x_pos), hw_f, z_f]))
        # 纳入"补齐底部"生成的点，使蒙皮贴合真实船底（而非一刀切到 y=0）
        bps = self.SurfaceGenerationData.get('BottomPoints')
        if bps is not None and len(bps):
            dense.append(np.asarray(bps, dtype=float))
        if not dense:
            messagebox.showerror('无数据', '未找到有效的曲线数据，无法生成点云。', parent=self.root)
            return
        dense_pts = np.unique(np.vstack(dense), axis=0)
        # 插值曲面 (X,Z)->Y，只使用右舷；linear + 加密点平滑度与 MATLAB natural 接近
        from scipy.interpolate import griddata
        Xp = dense_pts[:, 0]
        Yp = dense_pts[:, 1]
        Zp = dense_pts[:, 2]
        x_min, x_max = np.min(Xp), np.max(Xp)
        z_min, z_max = np.min(Zp), np.max(Zp)
        # 全精度数据网格：供 STL 导出等使用（生成一次，不直接用于交互显示）
        nx = max(40, min(120, round((x_max - x_min) / max(z_max - z_min, 1e-9) * 40)))
        nz = 80
        xlin = np.linspace(x_min, x_max, nx)
        zlin = np.linspace(z_min, z_max, nz)
        Xg, Zg = np.meshgrid(xlin, zlin)
        Yg = griddata((Xp, Zp), Yp, (Xg, Zg), method='linear')
        # 数据覆盖凸包外（如船首尾上缘）线性插值返回 NaN：
        # 填 0 使其在 _hull_mesh_with_bottom 的"有效半宽"判定中被剔除，
        # 不生成假几何——曲面在型线覆盖范围内自然收边（与 MATLAB NaN 处理等价）。
        Yg = np.nan_to_num(Yg, nan=0.0)
        # 存储全精度网格与面片（STL 导出用）
        # 用"边界曲面封底"：侧面按有效半宽构面，底部由边界→龙骨(y=0)直接闭合
        vertices_r, faces_r = self._hull_mesh_with_bottom(Xg, Yg, Zg)
        if vertices_r.shape[0] < 3:
            messagebox.showerror('生成失败', '插值曲面未能生成有效三角网格。', parent=self.root)
            return
        vertices_l = vertices_r.copy()
        vertices_l[:, 1] = -vertices_l[:, 1]
        faces_l = faces_r + vertices_r.shape[0]
        vertices = np.vstack([vertices_r, vertices_l])
        faces = np.vstack([faces_r, faces_l])
        self.SurfaceGenerationData['Vertices'] = vertices
        self.SurfaceGenerationData['Faces'] = faces
        self.SurfaceGenerationData['grid'] = (Xg, Yg, Zg)
        self.SurfaceGenerationData['display_faces'] = None  # 下方降采样后填写
        # 全精度网格推给 Qt SolidWorks 风格视窗（若已创建）
        if getattr(self, 'qt3d_host', None) is not None:
            self.push_qt_mesh(vertices, faces)
        # 显示用降采样网格：matplotlib 3D 是 CPU 软件渲染，面片数直接决定旋转流畅度
        k = self._display_mesh_step(nz, nx)
        dXg, dYg, dZg = Xg[::k, ::k], Yg[::k, ::k], Zg[::k, ::k]
        dVr, dFr = self._hull_mesh_with_bottom(dXg, dYg, dZg)
        dVl = dVr.copy()
        dVl[:, 1] = -dVl[:, 1]
        dFl = dFr + dVr.shape[0]
        dV = np.vstack([dVr, dVl])
        dF = np.vstack([dFr, dFl])
        self.SurfaceGenerationData['display_faces'] = int(len(dF))
        # 绘图
        p = self.plot_face_area
        p.clear()
        from mpl_toolkits.mplot3d.art3d import Poly3DCollection
        # antialiased=False 是 matplotlib 3D 最大的渲染加速项（软件光栅化）
        mesh = Poly3DCollection(dV[dF], facecolor=self.SurfaceColor, alpha=0.85,
                                edgecolor=(0.2, 0.2, 0.2) if self.WireframeMode == '高光边缘' else 'none',
                                linewidth=0.5, antialiased=False)
        p.ax.add_collection3d(mesh)
        # 保存 artist 引用：供"蒙皮颜色…"按钮在 matplotlib 回退视图下即时改色
        try:
            self._hull_mesh_artist = mesh
        except Exception:
            pass
        p.ax.set_xlabel('X (m)')
        p.ax.set_ylabel('Y (m)')
        p.ax.set_zlabel('Z (m)')
        p.ax.set_title('船体蒙皮（%s）' % self._mesh_quality_name())
        p.ax.view_init(elev=self.var_elevation.get(), azim=self.var_azimuth.get())
        p.ax.auto_scale_xyz([x_min, x_max], [-float(np.max(np.abs(Yp))), float(np.max(np.abs(Yp)))],
                            [z_min, z_max])
        p.set_true_box_aspect()
        p.refresh()
        self.log('船体蒙皮生成成功（质量: %s）。' % self._mesh_quality_name())
        messagebox.showinfo('成功', '船体蒙皮生成成功！可在"文件>导出"中导出 STL。', parent=self.root)

    def _mesh_quality_name(self):
        try:
            return self.var_mesh_quality.get() if hasattr(self, 'var_mesh_quality') else '标准'
        except Exception:
            return '标准'

    def _display_mesh_step(self, ny, nx):
        """按"蒙皮质量"档位计算显示网格的降采样步长 k。

        目标显示面片数（两舷合计）：流畅≈700、标准≈1500、精细≈3500。
        matplotlib 3D 为 CPU 软件渲染，面片越少旋转越流畅。
        """
        q = self._mesh_quality_name()
        target = {'流畅': 700, '标准': 1500, '精细': 3500}.get(q, 1500)
        nv = max(1, int(ny) * int(nx))
        # 显示面片≈4·nv/k²（两舷、每格两三角）；系数 3.2 使实际面片贴近目标档位
        k = int(math.sqrt(3.2 * nv / max(target, 100)))
        return max(1, min(k, min(ny, nx) // 2))

    def _hull_mesh_with_bottom(self, Xg, Yg, Zg, eps=0.005):
        """船体侧面三角网格 + 边界曲面封底。

        与 MATLAB 原版（生成蒙皮后底部开口）不同：这里直接把底面闭合。
        - 侧面：仅对"半宽 > eps"的网格单元构面，保持船体光滑；
        - 底面：逐列取最低的有效半宽点构成底部边界，用 边界→龙骨(y=0)
          的规则三角带直接闭合（边界曲面），替代 nan→0 的平底；
        - 相邻列底边界高度不同时（首尾阶梯），用 y=0 过渡三角补缝。
        返回 (vertices, faces)。
        """
        Xg = np.asarray(Xg, dtype=float)
        Yg = np.asarray(Yg, dtype=float)
        Zg = np.asarray(Zg, dtype=float)
        ny, nx = Xg.shape
        valid = Yg > eps
        verts = np.column_stack([Xg.ravel(), Yg.ravel(), Zg.ravel()])
        vid = np.arange(ny * nx).reshape(ny, nx)

        def add_pt(x, y, z):
            nonlocal verts
            verts = np.vstack([verts, [float(x), float(y), float(z)]])
            return len(verts) - 1

        faces = []
        # 1. 侧面：四角均有效
        for i in range(ny - 1):
            for j in range(nx - 1):
                if valid[i, j] and valid[i, j + 1] and valid[i + 1, j] and valid[i + 1, j + 1]:
                    a, b, c, d = vid[i, j], vid[i, j + 1], vid[i + 1, j], vid[i + 1, j + 1]
                    faces.append([a, b, c])
                    faces.append([b, d, c])
        # 2. 每列最低有效行（底部边界）
        bottom_row = np.full(nx, -1)
        for j in range(nx):
            rows = np.nonzero(valid[:, j])[0]
            if rows.size:
                bottom_row[j] = rows[0]
        cols = [j for j in range(nx) if bottom_row[j] >= 0]
        # 龙骨点 K_j = (x_j, 0, z_边界)
        keel = {}
        for j in cols:
            i = int(bottom_row[j])
            keel[j] = add_pt(Xg[i, j], 0.0, Zg[i, j])
        # 3. 封底三角带 + 相邻列阶梯过渡
        for a, b in zip(cols, cols[1:]):
            ia, ib = int(bottom_row[a]), int(bottom_row[b])
            ba, bb = vid[ia, a], vid[ib, b]
            ka, kb = keel[a], keel[b]
            faces.append([ba, bb, ka])
            faces.append([bb, kb, ka])
            i0, i1 = min(ia, ib), max(ia, ib)
            for i in range(i0, i1):
                # 列 a 在行 i / i+1 的角点（低于其底边界时投影到 y=0）
                pa_i = vid[i, a] if i >= ia else add_pt(Xg[i, a], 0.0, Zg[i, a])
                pa_i1 = vid[i + 1, a] if i + 1 >= ia else add_pt(Xg[i + 1, a], 0.0, Zg[i + 1, a])
                pb_i = vid[i, b] if i >= ib else add_pt(Xg[i, b], 0.0, Zg[i, b])
                pb_i1 = vid[i + 1, b] if i + 1 >= ib else add_pt(Xg[i + 1, b], 0.0, Zg[i + 1, b])
                faces.append([pa_i, pb_i, pa_i1])
                faces.append([pb_i, pb_i1, pa_i1])
        return verts, np.array(faces)

    def _grid_to_mesh(self, Xg, Yg, Zg):
        """规则网格转三角面片，返回 (vertices, faces)"""
        ny, nx = Xg.shape
        pts = []
        faces = []
        for i in range(ny):
            for j in range(nx):
                pts.append([Xg[i, j], Yg[i, j], Zg[i, j]])
        def idx(i, j):
            return i * nx + j
        for i in range(ny - 1):
            for j in range(nx - 1):
                a = idx(i, j)
                b = idx(i, j + 1)
                c = idx(i + 1, j)
                d = idx(i + 1, j + 1)
                faces.append([a, b, c])
                faces.append([b, d, c])
        return np.array(pts), np.array(faces)

    @_with_busy('正在导出 STL...')
    def export_stl(self):
        """导出船体蒙皮为 STL 文件"""
        if not self.SurfaceGenerationData or 'Vertices' not in self.SurfaceGenerationData \
                or 'Faces' not in self.SurfaceGenerationData:
            messagebox.showerror('导出失败', '未找到船体蒙皮数据，请先在 3D 曲面页点击"生成蒙皮"。',
                                 parent=self.root)
            return
        path = filedialog.asksaveasfilename(
            title='导出船体模型为 STL',
            filetypes=[('STL 文件', '*.stl')], defaultextension='.stl', parent=self.root)
        if not path:
            return
        try:
            verts = self.SurfaceGenerationData['Vertices']
            faces = self.SurfaceGenerationData['Faces']
            with open(path, 'w', encoding='ascii') as f:
                f.write('solid hull\n')
                for face in faces:
                    f.write('  facet normal 0 0 0\n    outer loop\n')
                    for i in face:
                        v = verts[i]
                        f.write('      vertex %.6e %.6e %.6e\n' % (v[0], v[1], v[2]))
                    f.write('    endloop\n  endfacet\n')
                f.write('endsolid hull\n')
            self.log('船体模型已成功导出为 STL 文件: %s' % path)
            messagebox.showinfo('导出成功', '船体模型已成功导出为 STL 文件。', parent=self.root)
        except Exception as e:
            messagebox.showerror('导出失败', '导出 STL 文件时出错: %s' % e, parent=self.root)

    def fill_bottom_points(self):
        """补齐底部点云（模仿 MATLAB Button_FillBottomPointsClicked）。

        对每个站号：取该站最低点 (minY, minZ)，若底部未闭合（minY≠0）则在
        Y=0 → minY 之间插值生成 3~8 个底部点（Z=minZ），从而把船底补到中线。
        同时更新镜像点云（AllPoints）、站号数组并重绘（原始红 / 底部蓝）。
        """
        sd = self.SurfaceGenerationData
        if not sd or 'HalfPoints' not in sd:
            messagebox.showerror('无数据', '尚未生成点云数据，请先点击"生成点云"。', parent=self.root)
            return
        hp = np.asarray(sd['HalfPoints'], dtype=float)
        half_stations = np.asarray(sd.get('HalfStations', []), dtype=float)
        if half_stations.size != hp.shape[0]:
            # 无站号数据时按 X 分组（同一纵向位置视为同一站）
            half_stations = hp[:, 0].copy()
        self.log('正在补齐底部点云...')
        unique_stations = np.unique(half_stations)
        bottom_parts = []
        st_parts = []
        for st in unique_stations:
            mask = half_stations == st
            pts = hp[mask]
            if len(pts) == 0:
                continue
            # 该站最低点（最小 Z）
            iz = int(np.argmin(pts[:, 2]))
            min_y = float(pts[iz, 1])
            min_z = float(pts[iz, 2])
            x_val = float(pts[iz, 0])
            # 最低点已在中线上（|Y|≤容差）→ 该站底部已闭合，无需补齐
            if abs(min_y) <= 0.001:
                continue
            # 根据距离选择点数（3~8，与 MATLAB 一致：min(8, max(3, round(|minY|*5)))）
            n = min(8, max(3, int(round(abs(min_y) * 5))))
            # Y 从 0（中线）线性插值到 minY；Z 取最低点高度
            fill_y = np.linspace(0.0, min_y, n)
            bottom_parts.append(np.column_stack([
                np.full(n, x_val), fill_y, np.full(n, min_z)]))
            st_parts.append(np.full(n, st))
        if not bottom_parts:
            self.log('底部点云已完整，无需补齐。')
            messagebox.showinfo('提示', '底部点云已完整，无需补齐。', parent=self.root)
            return
        bottom = np.vstack(bottom_parts)
        bottom_st = np.concatenate(st_parts)
        new_half = np.vstack([hp, bottom])
        new_st = np.concatenate([half_stations, bottom_st])
        # 去重（保持原顺序）
        _, uidx = np.unique(new_half, axis=0, return_index=True)
        uidx = np.sort(uidx)
        new_half = new_half[uidx]
        new_st = new_st[uidx]
        sd['HalfPoints'] = new_half
        sd['HalfStations'] = new_st
        sd['BottomPoints'] = bottom  # 半船底部补齐点
        # 镜像生成完整点云
        mirrored = new_half.copy()
        mirrored[:, 1] = -mirrored[:, 1]
        sd['AllPoints'] = np.unique(np.vstack([new_half, mirrored]), axis=0)
        sd['AllStations'] = np.concatenate([new_st, new_st])
        # 重绘：原始红 + 底部蓝（matplotlib 回退视图）
        self._redraw_pointcloud(bottom)
        # Qt SolidWorks 风格视窗立即同步：点云 + 型线底部轮廓（matplotlib 回退时无 Qt 静默）
        try:
            if getattr(self, 'qt3d_host', None) is not None:
                self._push_qt_pointcloud()
                self._push_qt_lines()
                self.qt3d_host.widget.update()
        except Exception:
            pass
        n_added = len(bottom) * 2
        self.log('底部点云补齐完成！新增 %d 个点（含镜像）。' % n_added)
        messagebox.showinfo('补齐成功',
                            '底部点云补齐完成！\n新增 %d 个点（含镜像）。' % n_added,
                            parent=self.root)

    def export_pointcloud(self):
        if not self.SurfaceGenerationData or 'HalfPoints' not in self.SurfaceGenerationData:
            messagebox.showerror('无数据', '尚未生成点云数据，请先点击"生成点云"。', parent=self.root)
            return
        path = filedialog.asksaveasfilename(
            title='导出点云数据',
            filetypes=[('CSV 文件', '*.csv'), ('Excel 文件', '*.xlsx'), ('文本文件', '*.txt')],
            defaultextension='.csv', parent=self.root)
        if not path:
            return
        hp = self.SurfaceGenerationData['HalfPoints']
        stations = self.SurfaceGenerationData.get('HalfStations', np.arange(len(hp)))
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.xlsx':
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(['站号', 'X_mm', 'Y_mm', 'Z_mm'])
                for i in range(len(hp)):
                    ws.append([stations[i], hp[i, 0] * 1000, hp[i, 1] * 1000, hp[i, 2] * 1000])
                wb.save(path)
            else:
                import csv
                sep = '\t' if ext == '.txt' else ','
                with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                    w = csv.writer(f, delimiter=sep)
                    w.writerow(['站号', 'X_mm', 'Y_mm', 'Z_mm'])
                    for i in range(len(hp)):
                        w.writerow([stations[i], '%.3f' % (hp[i, 0] * 1000),
                                    '%.3f' % (hp[i, 1] * 1000), '%.3f' % (hp[i, 2] * 1000)])
            messagebox.showinfo('导出成功', '成功导出 %d 个点。' % len(hp), parent=self.root)
        except Exception as e:
            messagebox.showerror('导出失败', str(e), parent=self.root)


def plt_cmap(n):
    """生成一组颜色"""
    import matplotlib.cm as cm
    cmap = cm.get_cmap('tab10')
    return [cmap(i % 10) for i in range(n)]
